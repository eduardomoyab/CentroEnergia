import os
import pandas as pd
import numpy as np
import time
import pickle as pkl
from queue import Queue
import copy as cp
import tkinter as tk
from tkinter import Tk, filedialog
import warnings
warnings.filterwarnings("ignore")
from functools import reduce
import openpyxl
from progress.bar import Bar, ChargingBar
import xlwings as xw
from openpyxl.styles import Color,PatternFill



nombre_err = ''
class Escenario_Base(object):
    def __init__(self, nombre_archivo: str, p_o_c: str, info_queue: Queue(), direction):
        global nombre_err
        self.archivo_excel = pd.ExcelFile(nombre_archivo)
        self.tipo = p_o_c
        try:
            nombre_err = 'Nombre del caso'
            self.nombre_escenario = pd.read_excel(nombre_archivo, "nombre_caso")["Nombre_Caso"][0]

            # Nuevo, se definen los set
            nombre_err = "'set_regiones'"
            self.regionMain = (pd.read_excel(self.archivo_excel, 'set_regiones')).to_numpy()
            nombre_err = "'set_ambito'"
            self.ambitoMain = (pd.read_excel(self.archivo_excel, 'set_ambito')).to_numpy()
            nombre_err = "'set_motorizacion'"
            self.motorizacionMain = (pd.read_excel(self.archivo_excel, 'set_motorizacion')).to_numpy()
            nombre_err = "'set_modos_pasajero'"
            self.modoPasajeroMain = (pd.read_excel(self.archivo_excel, 'set_modos_pasajero')).to_numpy()
            nombre_err = "'set_modos_carga'"
            self.modoCargaMain = (pd.read_excel(self.archivo_excel, 'set_modos_carga')).to_numpy()
            nombre_err = "'set_norma'"
            self.normaMain = (pd.read_excel(self.archivo_excel, 'set_norma')).to_numpy()
            nombre_err = "'set_gas'"
            self.gasMain = (pd.read_excel(self.archivo_excel, 'set_gas')).to_numpy()
            nombre_err = "'set_energeticos'"
            self.energeticosMain = (pd.read_excel(self.archivo_excel, 'set_energeticos')).to_numpy()
            nombre_err = "'set_tipos_vehiculos'"
            self.tiposVehiculosMain = (pd.read_excel(self.archivo_excel, 'set_tipos_vehiculos')).to_numpy()
            nombre_err = 'set_años'
            ages = (pd.read_excel(self.archivo_excel, 'set_años')).to_numpy()
            self.maxAge = str(np.max(ages))
            self.minAge = str(np.min(ages))  # 2060 y 1990
            nombre_err = 'set_modo_maritimo'
            self.modomarMain = (pd.read_excel(self.archivo_excel, 'set_modo_maritimo')).to_numpy()
            nombre_err = 'set_modo_aereo'
            self.modoaireMain = (pd.read_excel(self.archivo_excel, 'set_modo_aereo')).to_numpy()
        except:
            mensaje = "Error al ingresar valores de '", nombre_err, "' por favor revise que esté bien creada y nombrada la pagina en el archivo excel"
            info_queue.put(mensaje)
        else:
            try:
                print()
                if self.tipo == 'pasajero':
                    print("leyendo sets para pasajeros")
                    info_queue.put("leyendo sets para pasajeros")
                    nombre_err = "datos_pkm"
                    self.pkm_df = self.agregarDatosHojas(direction,info_queue, nombre_variable='pkm_modo',
                                                         nombre_hoja="datos_pkm")
                    nombre_err = "datos_participacion_modal"
                    
                    self.part_modal_df = self.agregarDatosHojas(direction, info_queue,
                                                                nombre_variable='participacion_modal',
                                                                nombre_hoja="datos_participacion_modal")
                    
                    nombre_err = "datos_tasa_ocupacion"
                    self.tasa_ocupacion_df = self.agregarDatosHojas(direction, info_queue,
                                                                    nombre_variable='tasa_ocupacion',
                                                                    nombre_hoja="datos_tasa_ocupacion")
                    nombre_err = "datos_poblacion"
                    self.poblacion_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='poblacion',
                                                               nombre_hoja="datos_driver_poblacion")
                    nombre_err = "datos_participacion_tecnologia"
                    self.part_tecnologica_df = self.agregarDatosHojas(direction, info_queue,
                                                                      nombre_variable='participacion_tecnologia',
                                                                      nombre_hoja="datos_participacion_tecnologia")
                    
                elif self.tipo == "carga":
                    print("leyendo sets para carga")
                    info_queue.put("leyendo sets para carga")
                    nombre_err = "datos_tkm"
                    self.tkm_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='tkm_modo',
                                                         nombre_hoja="datos_tkm")
                    nombre_err = "datos_participacion_modal_carga"
                    self.part_modal_df = self.agregarDatosHojas(direction, info_queue,
                                                                nombre_variable='participacion_modal',
                                                                nombre_hoja="datos_participacion_modal_carga")
                    nombre_err = "datos_tasa_carga"
                    self.tasa_carga_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='tasa_carga',
                                                                nombre_hoja="datos_tasa_carga")
                    nombre_err = "datos_part_tecnologia_carga"
                    self.part_tecnologica_df = self.agregarDatosHojas(direction, info_queue,
                                                                      nombre_variable='participacion_tecnologia',
                                                                      nombre_hoja="datos_part_tecnologia_carga")

                elif self.tipo == "pasajero-aereo":
                    print("leyendo sets para pasajero aereo")
                    info_queue.put("leyendo sets para pasajero aereo")
                    nombre_err = "datos_pkm_aereo"
                    self.pkm_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='pkm_modo',
                                                         nombre_hoja="datos_pkm_aereo")
                    nombre_err = "datos_ajuste_aereo"
                    self.pkm_ajuste_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='pkm_ajuste',
                                                                nombre_hoja="datos_factor_ajuste_aereo")
                    nombre_err = "datos_tasa_ocupacion_aereo"
                    self.tasa_ocupacion_df = self.agregarDatosHojas(direction, info_queue,
                                                                    nombre_variable='tasa_ocupacion',
                                                                    nombre_hoja="datos_tasa_ocupacion_aereo")
                    nombre_err = "datos_participacion_tecno_aereo"
                    self.part_tecnologica_df = self.agregarDatosHojas(direction, info_queue,
                                                                      nombre_variable='participacion_tecnologia',
                                                                      nombre_hoja="datos_participacion_tecno_aereo")

                elif self.tipo == "carga-maritimo":
                    print("leyendo sets para carga marítimo")
                    info_queue.put("leyendo sets para carga marítimo")
                    nombre_err = "datos_tkm_marítimo"
                    self.tkm_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='tkm_modo',
                                                         nombre_hoja="datos_tkm_marítimo")
                    nombre_err = "datos_ajuste_marítimo"
                    self.tkm_ajuste_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='tkm_ajuste',
                                                                nombre_hoja="datos_factor_ajuste_marítimo")

                    nombre_err = "datos_tasa_carga_marítimo"
                    self.tasa_carga_df = self.agregarDatosHojas(direction, info_queue, nombre_variable='tasa_carga',
                                                                nombre_hoja="datos_tasa_carga_marítimo")
                    nombre_err = "datos_participacion_tecno_marit"
                    self.part_tecnologica_df = self.agregarDatosHojas(direction, info_queue,
                                                                      nombre_variable='participacion_tecnologia',
                                                                      nombre_hoja="datos_participacion_tecno_marit")
                
            except:
                pass
    #Funcion para agregar valores de las hojas set
    def _agregar_datos(self, nombre_variable: str, nombre_hoja: str, constantes=False) -> pd.DataFrame:
        datos = pd.read_excel(self.archivo_excel, nombre_hoja)
        datos = self.cambiar_estructura(datos, constantes)
        datos = datos.rename(columns={'Valor': nombre_variable})
        return datos

    #   Funcion que crea un diccionario que va enlazando todas las columnas de años de las hojas del excel
    #   con el tipo de dato numérico Float
    def createDicc(self, nombre_hoja: str) -> dict:
        datos = pd.read_excel(self.archivo_excel, nombre_hoja)
        number = datos.columns.get_loc("Unidad")
        dicc = {}
        for i in datos.iloc[:, number + 1:]:
            dicc[i] = float
        return dicc

    @staticmethod
    def cambiar_estructura(df_archivo, constantes=False):
        # Se borra la columna unidad ya que no tiene importancia en el codigo
        try:
            df_archivo = df_archivo.drop(columns="Unidad")
        except:
            pass
        lista_columnas = df_archivo.columns.values.tolist()
        # Caso en que hay columnas con años
        if not constantes:
            lista_columnas_transformada = []
            for columna in lista_columnas:
                try:
                    lista_columnas_transformada.append(int(columna))
                except:
                    pass
            indices = [x for x in lista_columnas if x not in set(lista_columnas_transformada)]

            new_names = [(i, "Valor" + str(i)) for i in df_archivo.columns if i not in indices]
            df_archivo.rename(columns=dict(new_names), inplace=True)
            nuevo_df = pd.wide_to_long(df_archivo, stubnames="Valor",
                                       i=indices, j='Año')
        # Caso en que hay constantes
        # Se considera que la ultima columna contendra la constante y las otras se transforma en indice
        else:
            nuevo_df = df_archivo.set_index(lista_columnas[:-1])
        return nuevo_df.sort_index()

    
    #   Funcion para agregar valores de las hojas con datos, además se encarga de llamar a la funcion
    #   Que verifica la correctitu de los datos ingresados, esto para ahorrar tiempo
    def agregarDatosHojas(self, direction, info_queue: Queue(), nombre_variable: str, nombre_hoja: str) -> pd.DataFrame:
        dicc = self.createDicc(nombre_hoja)
        datos = pd.read_excel(self.archivo_excel, nombre_hoja, dtype=dicc)

        #self._verificador_columnas_set(direction, datos, nombre_hoja, info_queue)
        #ErrConflic = open(direction + "Errores_Conflictivos.txt", 'a+')
        #self._anho_columnas(datos, nombre_hoja, info_queue, ErrConflic)
        #ErrConflic.close()

        datos = self.cambiar_estructura(datos, constantes=False)
        datos = datos.rename(columns={'Valor': nombre_variable})
        return datos

    #   Busca en una columna con un valor 'object', el dato que corrompe la estructura. Indicandole al usuario
    #   La posicion del dato conflictivo
    def buscarEnColumna(self, nombre_hoja: str, info_queue: Queue(), TexErr) -> None:
        datos = pd.read_excel(self.archivo_excel, nombre_hoja)
        number = datos.columns.get_loc("Unidad")

        for column in datos.iloc[:, number + 1:]:
            if datos[column].dtype != float:
                for i in range(datos.shape[0]):
                    try:
                        float(datos.iloc[i][column])
                    except:
                        j = i + 2
                        mensaje = "ERROR DE INGRESO: El valor numero " + str(j) + " de la columna " + str(
                            column) + " en la hoja " + nombre_hoja + " no es numerico" + '\n'
                        TexErr.writelines(mensaje)

    #   Funcion que verifica los valores de las columnas que representen los valores de los sets
    def _verificador_columnas_set(self, direction, datos: pd.DataFrame, nombre_hoja: str, info_queue: Queue()) -> None:
        iter = datos.columns.get_loc("Unidad")
        for i in datos.iloc[:, :iter]:
            if i == 'Region':
                self._alarma(direction, datos["Region"].unique(), nombre_hoja, self.regionMain)
            elif i == 'Ambito':
                self._alarma(direction, datos["Ambito"].unique(), nombre_hoja, self.ambitoMain)
            elif i == 'Modo':
                if self.tipo == 'pasajero':
                    self._alarma(direction, datos["Modo"].unique(), nombre_hoja, self.modoPasajeroMain)
                else:  # self.tipo == 'carga'
                    self._alarma(direction, datos["Modo"].unique(), nombre_hoja, self.modoCargaMain)
            elif i == 'Modo_maritimo':
                self._alarma(direction, datos["Modo_Maritimo"].unique(), nombre_hoja, self.modomarMain)
            elif i == 'Modo_Aereo':
                self._alarma(direction, datos["Modo_Aereo"].unique(), nombre_hoja, self.modoaireMain)
            elif i == 'Motorizacion':
                self._alarma(direction, datos["Motorizacion"].unique(), nombre_hoja, self.motorizacionMain)
            elif i == 'Norma':
                self._alarma(direction, datos["Norma"].unique(), nombre_hoja, self.normaMain)
            elif i == 'Energético':
                self._alarma(direction, datos["Energético"].unique(), nombre_hoja, self.energeticosMain)
            elif i == 'Gas':
                self._alarma(direction, datos["Gas"].unique(), nombre_hoja, self.gasMain)
            else:
                self.verAdicionalSet(i, datos[i].unique(), nombre_hoja, direction)

    def verAdicionalSet(self, i, datos_unicos, nombre_hoja, direction):
        aux = 'set_' + i
        try:
            columnMain = (pd.read_excel(self.archivo_excel, aux)).to_numpy()
        except:
            ErrConflic = open(f"{direction}Errores_Conflictivos.txt", 'a+')
            mensaje = f"El nombre de columnam {i} de la hoja {nombre_hoja} no tiene un set asignado correctamente \n"
            ErrConflic.writelines(mensaje)
            ErrConflic.close()
        else:
            self._alarma(direction, datos_unicos, nombre_hoja, columnMain)

    def _alarma(self, direction, listaDatos: list, nombre_hoja: str, listaMain: list) -> None:
        ErrAusencia = open(f"{direction}Errores_Ausencia.txt", 'a+')
        self.verificarAusencia(listaDatos, nombre_hoja, listaMain, ErrAusencia)
        ErrAusencia.close()
        ErrConflic = open(f"{direction}Errores_Conflictivos.txt", 'a+')
        self.verificarConflictivo(listaDatos, nombre_hoja, listaMain, ErrConflic)
        ErrConflic.close()

    #   Funcion que se encarga de indicarle al usuario la existencia de un valor indicado
    #   en el set pero que no está siendo utilizado
    def verificarAusencia(self, listaDatos: list, nombre_hoja: str, listaMain: list, TexErr) -> None:
        for i in listaMain:
            if i not in listaDatos:
                mensaje = f"AUSENCIA DATO: No se encontraron datos de {i} en la hoja {nombre_hoja}\n"
                TexErr.writelines(mensaje)

    #   Funcion que se encarga de indicarle al usuario que existe un dato ingresado que no
    #   está contenido en los sets
    def verificarConflictivo(self, listaDatos: list, nombre_hoja: str, listaMain: list, TexErr) -> None:
        for i in listaDatos:
            if i not in listaMain:
                try:
                    mensaje = f"DATO CONFLICTIVO: El valor {i} en la hoja {nombre_hoja} no corresponde a ningun valor " \
                              f"de los set establecidos \n "
                except:
                    mensaje = f"DATO CONFLICTIVO: Existe un valor vacío dentro de la hoja {nombre_hoja}\n"
                TexErr.writelines(mensaje)

                #   Funcion que verifica que las columnas de años ingresados estén dentro de los parámetros

    #   indicados en el set años
    def _anho_columnas(self, datos: pd.DataFrame, nombre_hoja: str, info_queue: Queue(), TexErr) -> None:
        number = datos.columns.get_loc("Unidad")

        for i in datos.iloc[:, number + 1:]:
            i = str(i)
            if i > self.maxAge:
                mensaje = "El año '" + i + "' de la hoja '" + nombre_hoja + "' es mayor al máximo indicado en el set" + '\n'
                TexErr.writelines(mensaje)
            elif i < self.minAge:
                mensaje = "El año '" + i + "' de la hoja '" + nombre_hoja + "' es menor al mínimo indicado en el set" + '\n'
                TexErr.writelines(mensaje)

    def calcular_pkm_tecnologia(self, opc1, opc2, opc3, par_tec) -> pd.DataFrame:
        if self.tipo == "pasajero":
            Resultados = self.pkm_df.join(self.part_modal_df, how="inner")
            if opc1 == "exógena" and opc2 == "exógena":
                Resultados = Resultados.join(self.part_tecnologica_df, how="inner")
            if opc1 == "endógena" or opc2 == "endógena":
                Resultados = Resultados.join(par_tec, how="inner")

            # Calculo de PKM por tecnologia
            pkm = Resultados['pkm_modo']
            participacion_modal = Resultados['participacion_modal']
            participacion_tecnologia = Resultados['participacion_tecnologia']
            # VKM_Tecnologia (km) = pkm (km) * participacion_modal (%) * participacion_tecnologia (%)
            Resultados['PKM_Tecnologia'] = pkm * participacion_modal * participacion_tecnologia
            return Resultados.fillna(0)

        elif self.tipo == "pasajero-aereo":
            Resultados = self.pkm_df.join(self.pkm_ajuste_df, how="inner")
            Resultados = Resultados.join(self.part_tecnologica_df, how="inner")

            # Calculo de VKM por tecnologia
            pkm = Resultados['pkm_modo']
            participacion_tecnologia = Resultados['participacion_tecnologia']
            pkm_ajuste = Resultados['pkm_ajuste']

            # PKM_Tecnologia (km) = pkm (km) * participacion_tecnologia (%)
            Resultados['PKM_Tecnologia'] = pkm * pkm_ajuste * participacion_tecnologia
            return Resultados.fillna(0)


    def calcular_tkm_tecnologia(self, opc1, opc2, opc3, par_tec) -> pd.DataFrame:
        if self.tipo == "carga":
            Resultados = self.tkm_df.join(self.part_modal_df, how="inner")

            if opc3 == "exógena":
                Resultados = Resultados.join(self.part_tecnologica_df, how="inner")
            elif opc3 == "endógena":
                Resultados = Resultados.join(par_tec, how="inner")

            # Calculo de VKM
            tkm = Resultados['tkm_modo']
            participacion_modal = Resultados['participacion_modal']
            participacion_tecnologia = Resultados['participacion_tecnologia']

            # TKM_Tecnologia (tkm) = tkm (ton-km) * participacion_modal (%)  * participacion_tecnologia (%)
            Resultados['TKM_Tecnologia'] = tkm * participacion_modal * participacion_tecnologia
            return Resultados.fillna(0)
        elif self.tipo == "carga-maritimo":
            Resultados = self.tkm_df.join(self.tkm_ajuste_df, how="inner")
            Resultados = Resultados.join(self.part_tecnologica_df, how="inner")

            # Calculo de VKM
            tkm = Resultados['tkm_modo']
            participacion_tecnologia = Resultados['participacion_tecnologia']
            tkm_ajuste = Resultados['tkm_ajuste']

            # VKM_Tecnologia (km) = tkm (ton-km) / tasa_carga (ton/veh) * participacion_tecnologia (%)
            Resultados['TKM_Tecnologia'] = tkm * tkm_ajuste * participacion_tecnologia
            return Resultados.fillna(0)


    def calcular_vkm_tecnologia(self, opc1, opc2, opc3, par_tec) -> pd.DataFrame:
        if self.tipo == "pasajero":
            resultados = self.pkm_df.join(self.part_modal_df, how="inner")
            resultados = resultados.join(self.tasa_ocupacion_df, how="inner")
            # Resultados = Resultados.join(self.part_tecnologica_df, how="inner")

            if opc1 == "exógena" and opc2 == "exógena":
                resultados = resultados.join(self.part_tecnologica_df, how="inner")
            elif opc1 == "endógena" or opc2 == "endógena":
                resultados = resultados.join(par_tec, how="inner")

            # Calculo de VKM por tecnologia
            pkm = resultados['pkm_modo']
            participacion_modal = resultados['participacion_modal']
            tasa_ocupacion = resultados['tasa_ocupacion']
            participacion_tecnologia = resultados['participacion_tecnologia']

            # VKM_Tecnologia (km) = pkm (km) * participacion_modal (%) / tasa_ocupacion (per/veh) *
            # participacion_tecnologia (%)
            resultados['VKM_Tecnologia'] = pkm * participacion_modal / tasa_ocupacion * participacion_tecnologia
            return resultados.fillna(0)
        elif self.tipo == "carga":
            resultados = self.tkm_df.join(self.part_modal_df, how="inner")
            resultados = resultados.join(self.tasa_carga_df, how="inner")
            # Resultados = Resultados.join(self.part_tecnologica_df, how="inner")

            if opc3 == "exógena":
                resultados = resultados.join(self.part_tecnologica_df, how="inner")
            elif opc3 == "endógena":
                resultados = resultados.join(par_tec, how="inner")

            # Calculo de VKM
            tkm = resultados['tkm_modo']
            participacion_modal = resultados['participacion_modal']
            tasa_carga = resultados['tasa_carga']
            participacion_tecnologia = resultados['participacion_tecnologia']

            # VKM_Tecnologia (km) = tkm (ton-km) * participacion_modal (%) / tasa_carga (ton/veh) *
            # participacion_tecnologia (%)
            resultados['VKM_Tecnologia'] = tkm * participacion_modal / tasa_carga * participacion_tecnologia
            return resultados.fillna(0)
        elif self.tipo == "pasajero-aereo":
            resultados = self.pkm_df.join(self.tasa_ocupacion_df, how="inner")
            resultados = resultados.join(self.part_tecnologica_df, how="inner")
            resultados = resultados.join(self.pkm_ajuste_df, how="inner")

            # Calculo de VKM por tecnologia
            pkm = resultados['pkm_modo']
            tasa_ocupacion = resultados['tasa_ocupacion']
            participacion_tecnologia = resultados['participacion_tecnologia']
            pkm_ajuste = resultados['pkm_ajuste']

            # VKM_Tecnologia (km) = pkm (km) / tasa_ocupacion (per/veh) * participacion_tecnologia (%)
            resultados['VKM_Tecnologia'] = pkm * pkm_ajuste / tasa_ocupacion * participacion_tecnologia
            return resultados.fillna(0)
        elif self.tipo == "carga-maritimo":
            resultados = self.tkm_df.join(self.tasa_carga_df, how="inner")
            resultados = resultados.join(self.part_tecnologica_df, how="inner")
            resultados = resultados.join(self.tkm_ajuste_df, how="inner")

            # Calculo de VKM
            tkm = resultados['tkm_modo']
            tasa_carga = resultados['tasa_carga']
            participacion_tecnologia = resultados['participacion_tecnologia']
            tkm_ajuste = resultados['tkm_ajuste']

            # VKM_Tecnologia (km) = tkm (ton-km) / tasa_carga (ton/veh) * participacion_tecnologia (%)
            resultados['VKM_Tecnologia'] = tkm * tkm_ajuste / tasa_carga * participacion_tecnologia
            return resultados.fillna(0)

def calculos_base(escenario: Escenario_Base, direction: str, op1, op2, op3, opcion4, opcion5):
    tipo = escenario.tipo
    part_tec_base = []
    
    print("Calculo de PKM por tecnologia")
    if tipo == "pasajero" or tipo == "pasajero-aereo":
        
        PKM_Tecnologia = escenario.calcular_pkm_tecnologia(op1, op2, op3, part_tec_base)

    print("Calculo de TKM por tecnologia")
    if tipo == "carga" or tipo == "carga-maritimo":
        TKM_Tecnologia = escenario.calcular_tkm_tecnologia(op1, op2, op3, part_tec_base)


    print("Calculo de VKM por tecnologia")
    VKM_Tecnologia = escenario.calcular_vkm_tecnologia(op1, op2, op3, part_tec_base)
    # VKM_Tecnologia = escenario.calcular_vkm_tecnologia()
    # VKM_Tecnologia.to_csv(direction +"solucion_" + tipo + "_vkm_tec_" + escenario.nombre_escenario + ".csv", encoding='utf-8-sig',sep=opcion4,decimal=opcion5)
    # print(VKM_Tecnologia)

# Armando los dataframes que tendran todos los escenarios
    if tipo == "pasajero":
        global PKM_Tecnologia_todos_pas
        PKM_Tecnologia_todos_pas = pd.concat({escenario.nombre_escenario: PKM_Tecnologia}, names=['Escenario'])
        global VKM_Tecnologia_todos_pas
        VKM_Tecnologia_todos_pas = pd.concat({escenario.nombre_escenario: VKM_Tecnologia}, names=['Escenario'])
    
    elif tipo == "carga":
        global TKM_Tecnologia_todos_carga
        TKM_Tecnologia_todos_carga = pd.concat({escenario.nombre_escenario: TKM_Tecnologia}, names=['Escenario'])
        global VKM_Tecnologia_todos_carga
        VKM_Tecnologia_todos_carga = pd.concat({escenario.nombre_escenario: VKM_Tecnologia}, names=['Escenario'])
        
    elif tipo == "pasajero-aereo":
        global PKM_Tecnologia_todos_pas_aereo
        PKM_Tecnologia_todos_pas_aereo = pd.concat({escenario.nombre_escenario: PKM_Tecnologia}, names=['Escenario'])
        global VKM_Tecnologia_todos_pas_aereo
        VKM_Tecnologia_todos_pas_aereo = pd.concat({escenario.nombre_escenario: VKM_Tecnologia}, names=['Escenario'])

    elif tipo == "carga-maritimo":
        global TKM_Tecnologia_todos_carga_maritimo
        TKM_Tecnologia_todos_carga_maritimo = pd.concat({escenario.nombre_escenario: TKM_Tecnologia},
                                                        names=['Escenario'])
        global VKM_Tecnologia_todos_carga_maritimo
        VKM_Tecnologia_todos_carga_maritimo = pd.concat({escenario.nombre_escenario: VKM_Tecnologia},
                                                        names=['Escenario'])
        
# Busca la columna años y agrega una columna con los valores de la unidad
def mostrarUnidad(Datos, unidad, direction, nombre, opcion4, opcion5):
    # Datos=pd.DataFrame({'email':datos.index, 'list':datos.values})
    try:
        datos = Datos.to_frame()
    except:
        datos = Datos.copy()
    datos['Unidad'] = unidad
    # number=datos.columns.get_loc("Año")
    # datos.insert(0, 'Unidad', s.values)

    datos.to_csv(direction+nombre + '.csv', encoding='utf-8-sig', sep=opcion4, decimal=opcion5)

# Función para agregar columna, recive la suma ya realizada y se encarga de agregarla al dataframe
# Retorna el nuevo dataframe
def addSum(dataframe, nombre, new_row, unidad, divisor):
    new_row = new_row.transpose()
    new_row.iloc[0, 0] = 0
    new_row.iloc[0, 1] = 0
    new_row = new_row.div(divisor, fill_value=float)
    new_row.iloc[0, 0] = nombre
    new_row.iloc[0, 1] = unidad
    return pd.concat([dataframe, new_row], axis=0, ignore_index=True)


# Crea un dicionario con las posiciones iniciales de las filas
def createDict(nombres_filas):
    dicc = {}
    it = 0
    for i in nombres_filas:
        dicc[i] = it
        it += 1
    return dicc

def main(ruta, caso_base, escenarios={}, progress_queue=Queue(), info_queue=Queue()):
    progress_queue.put(1)
    print("Obteniendo datos de sector transporte")
    start_time = time.time()
    
    # Auxiliar para manejo de direcciones
    dir = pd.read_excel(caso_base, 'Config')
    for i in escenarios:
        print(escenarios[i])
    
    direction = ruta
    direction += '\\'
    #direction=""

    '''ErrAusencia = open(direction + "Errores_Ausencia.txt", 'w')
        ErrAusencia.truncate(0)
        ErrAusencia.close()
        ErrConflic = open(direction + "Errores_Conflictivos.txt", 'w')
        ErrConflic.truncate(0)
        ErrConflic.close()'''

    # definir opcion exogena o endogena
    opcion1 = dir.iloc[3]['Respuesta']  # VLP
    opcion2 = dir.iloc[4]['Respuesta']  # VLC
    opcion3 = dir.iloc[5]['Respuesta']  # Camion
    opcion4 = dir.iloc[7]['Respuesta']  # Separador
    if opcion4 == 0:
        opcion4 = ' '
    opcion5 = dir.iloc[8]['Respuesta']  # Decimal

    try:
        
        # Proyecciones transporte pasajeros Escenario base
        progress_queue.put(5)
        info_queue.put('Cargando datos Pasajero')
        Escenario_pasajero = Escenario_Base(caso_base, "pasajero", info_queue, direction)
        progress_queue.put(10)
        info_queue.put('Calculando información Pasajero')
        calculos_base(Escenario_pasajero, direction, opcion1, opcion2, opcion3, opcion4, opcion5)
        
        
        # Proyecciones transporte carga Escenario base
        progress_queue.put(15)
        info_queue.put('Cargando datos Carga')
        Escenario_carga = Escenario_Base(caso_base, "carga", info_queue, direction)

        progress_queue.put(20)
        info_queue.put('Calculando información Carga')
        calculos_base(Escenario_carga, direction, opcion1, opcion2, opcion3, opcion4, opcion5)
        
        # Proyecciones transporte pasajeros en modo aereo Escenario base
        progress_queue.put(25)
        info_queue.put('Cargando datos Pasajero-Aéreo')
        Escenario_pasajero_aereo = Escenario_Base(caso_base, "pasajero-aereo", info_queue, direction)

        progress_queue.put(30)
        info_queue.put('Calculando información Pasajero-Aéreo')
        calculos_base(Escenario_pasajero_aereo, direction, opcion1, opcion2, opcion3, opcion4, opcion5)

        # Proyecciones transporte carga maritimo Escenario base
        progress_queue.put(35)
        info_queue.put('Cargando datos Carga-Marítimo')
        Escenario_carga_maritimo = Escenario_Base(caso_base, "carga-maritimo", info_queue, direction)

        progress_queue.put(40)
        info_queue.put('Calculando información Carga-Marítimo')
        calculos_base(Escenario_carga_maritimo, direction, opcion1, opcion2, opcion3, opcion4, opcion5)

    except:
        mensaje = "ERRROR AL IMPORTAR LOS DATOS DE '", nombre_err, "', verifique los .txt generados"
        info_queue.put(mensaje)
        return
    
    global PKM_Tecnologia_todos_pas
    global PKM_Tecnologia_todos_pas_aereo
    global PKM_Tecnologia_todos

    global TKM_Tecnologia_todos_carga
    global TKM_Tecnologia_todos_carga_maritimo
    global TKM_Tecnologia_todos

    global VKM_Tecnologia_todos_pas
    global VKM_Tecnologia_todos_carga
    
    # Pasando a csv todos los archivos con los escenarios juntos
    print("")
    print("Creando csv agrupados por escenario...")

    progress_queue.put(85)
    info_queue.put('Traspasando resultados a csv')
    
    # Pasando a csv todos los archivos con los datos de pasajeros y carga juntos
    # Pasajero vial
    # Se agrega columna "Pasajero" para posteriormente integrar todas las salidas en un unico archivo
    PKM_Tecnologia_todos_pas = pd.concat({"Pasajero": PKM_Tecnologia_todos_pas}, names=['Tipo'])
    VKM_Tecnologia_todos_pas = pd.concat({"Pasajero": VKM_Tecnologia_todos_pas}, names=['Tipo'])
    
    # Carga vial
    # Se agrega columna "Carga" para posteriormente integrar todas las salidas en un unico archivo
    TKM_Tecnologia_todos_carga = pd.concat({"Carga": TKM_Tecnologia_todos_carga}, names=['Tipo'])
    VKM_Tecnologia_todos_carga = pd.concat({"Carga": VKM_Tecnologia_todos_carga}, names=['Tipo'])

    # Pasajeo aereo
    # Se agrega columna "Pasajero-Aereo" para posteriormente integrar todas las salidas en un unico archivo
    PKM_Tecnologia_todos_pas_aereo = pd.concat({"Pasajero-Aereo": PKM_Tecnologia_todos_pas_aereo}, names=['Tipo'])

    # Carga maritimo
    # Se agrega columna "Carga-Maritimo" para posteriormente integrar todas las salidas en un unico archivo
    TKM_Tecnologia_todos_carga_maritimo = pd.concat({"Carga-Maritimo": TKM_Tecnologia_todos_carga_maritimo},
                                                    names=['Tipo'])
    
    
    
    
    # Aqui empiezo a concatenar todos los archivos anteriores para exportar las salidas en un unico archivo

    # Se exportan PKM con los datos de todos los escenarios
    PKM_Tecnologia_todos = pd.concat(
        [PKM_Tecnologia_todos_pas.reset_index(), PKM_Tecnologia_todos_pas_aereo.reset_index()])
    PKM_Tecnologia_todos.set_index(['Tipo'], inplace=True)
    mostrarUnidad(PKM_Tecnologia_todos, 'PKM', direction, "PKM_Tecnologia_todos", opcion4, opcion5)

    # Se exportan TKM con los datos de todos los escenarios
    TKM_Tecnologia_todos = pd.concat(
        [TKM_Tecnologia_todos_carga.reset_index(), TKM_Tecnologia_todos_carga_maritimo.reset_index()])
    TKM_Tecnologia_todos.set_index(['Tipo'], inplace=True)
    mostrarUnidad(TKM_Tecnologia_todos, 'TKM', direction, "TKM_Tecnologia_todos", opcion4, opcion5)

    # Se exportan VKM con los datos de todos los escenarios
    VKM_Tecnologia_todos = pd.concat([VKM_Tecnologia_todos_pas.reset_index(), VKM_Tecnologia_todos_carga.reset_index()])
    VKM_Tecnologia_todos.set_index(['Tipo'], inplace=True)
    mostrarUnidad(VKM_Tecnologia_todos, 'km', direction, "VKM_Tecnologia_todos", opcion4, opcion5)
    
    print("")
    print("tiempo ejecucion:")
    print("--- %s segundos ---" % (time.time() - start_time))

    progress_queue.put(100)
    info_queue.put('Ejecución finalizada con éxito')

#Estructura para realizar joins
def cambiar_estructura(df_archivo, nombre_variable:str,constantes=False):
    # Se borra la columna unidad ya que no tiene importancia en el codigo
    try:
        df_archivo = df_archivo.drop(columns="Unidad")
    except:
        pass
    lista_columnas = df_archivo.columns.values.tolist()
    # Caso en que hay columnas con años
    if not constantes:
        lista_columnas_transformada = []
        for columna in lista_columnas:
            try:
                lista_columnas_transformada.append(int(columna))
            except:
                pass
        indices = [x for x in lista_columnas if x not in set(lista_columnas_transformada)]

        new_names = [(i, "Valor" + str(i)) for i in df_archivo.columns if i not in indices]
        df_archivo.rename(columns=dict(new_names), inplace=True)
        nuevo_df = pd.wide_to_long(df_archivo, stubnames="Valor",
                                    i=indices, j='Año')
    # Caso en que hay constantes
    # Se considera que la ultima columna contendra la constante y las otras se transforma en indice
    else:
        nuevo_df = df_archivo.set_index(lista_columnas[:-1])
    
    nuevo_df = nuevo_df.rename(columns={'Valor': nombre_variable})
    return nuevo_df.sort_index()


#Transformamos los datos de entrada para realizar el cálculo
def transformar_entrada(path,obs):
    entrada = pd.read_excel(path,sheet_name='datos_entrada').dropna(subset=[2017]).drop(columns='Fuente')
    entrada.drop(columns='Unidad',inplace=True)
    entrada.set_index('Variable',inplace=True)

    proyec  = pd.read_excel(obs,sheet_name='Proyecciones demanda')
    proyec = proyec.filter(regex='^(?!Unnamed)')
    proyec = proyec.iloc[:, :-1]
    proyec.dropna(subset=['Año'],inplace=True)
    if 'km_recorridos' in proyec.columns:
        proyec.rename(columns={'km_recorridos': 'km_livianos'}, inplace=True)
    if 'km_recorridos_taxi' in proyec.columns:
        proyec.rename(columns={'km_recorridos_taxi': 'km_taxi'}, inplace=True)



    columnas_interes = proyec.columns[proyec.columns.get_loc('km_livianos'):proyec.columns.get_loc('tkm_ferroviario')+1]

    proyec = proyec.drop(columns=columnas_interes)
    proyec.set_index('Año',inplace=True)

    data = pd.DataFrame(index=entrada.columns, columns=proyec.columns)

    for año in entrada.columns:
        if año < 2017:
            pass
        else:
            for var in list(entrada.index):
                if var not in proyec.columns and var != 'Tipo cambio dólar' and var !='salitre':
                    pass
                else:
                    data.loc[año,var] = entrada.loc[var][año]
    data.index.name = 'Año'
    return data



def generar_salidas(path,com_aux=None):
    com_aux=dic_comercial
    datos = pd.read_excel(path,sheet_name='datos_proyeccion_demanda')
    wb = openpyxl.load_workbook(path)

    light_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    color = Color(rgb='002060')

    
    #Driver
    salida_driver = datos[(datos.Variable == 'PIB') | (datos.Variable == 'PIB (mill USD)') | (datos.Variable == 'tasa PIB') | (datos.Variable == 'Población') | (datos.Variable == 'Viviendas') | (datos.Variable == 'HAB/viv') | (datos.Variable == 'PIB per cápita (usd/hab)') |
    (datos.Variable == 'Tasa Población') | (datos.Variable == 'Tipo cambio dólar') | (datos.Variable == 'Tipo cambio euro')]

    pivote_driver = salida_driver.pivot_table(values='Valor', index=['Variable', 'Unidad'], columns='Año')
    pivote_driver.reset_index(inplace=True)
    pivote_driver.rename_axis(None, axis=1, inplace=True)


    order_list = ["tasa PIB", "Tipo cambio dólar", "Tipo cambio euro", "PIB", "PIB (mill USD)", "Tasa Población", "Población", "Viviendas", "HAB/viv","PIB per cápita (usd/hab)"]
    pivote_driver['orden'] = pivote_driver['Variable'].map(lambda x: order_list.index(x))
    pivote_driver = pivote_driver.sort_values(by='orden')
    pivote_driver.reset_index(inplace=True)
    pivote_driver.drop(columns=['orden','index'],inplace=True)

    sheet_name = 'salida_proyeccion_driver'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color

    for j, header in enumerate(pivote_driver.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_driver.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    #Energias
    ##Energia comercial
    salidas_comercial = datos[(datos.Variable == 'energia_comercial') | (datos['Variable'].str.endswith('_COM'))]

    pivote_comercial = salidas_comercial.pivot_table(values='Valor', index=['Variable', 'Unidad'], columns='Año')
    pivote_comercial.reset_index(inplace=True)
    pivote_comercial.rename_axis(None, axis=1, inplace=True)
    pivote_comercial['_COM'] = pivote_comercial.Variable.str.endswith('_COM')
    pivote_comercial.sort_values(by=['_COM'], ascending=True, inplace=True)
    pivote_comercial.reset_index(inplace=True)
    pivote_comercial.drop(columns=['_COM','index'],inplace=True)

    merge_df = pivote_comercial.merge(com_aux.reset_index(), how='left', left_on='Variable', right_on='Nombre_demanda')

    # seleccionar solo las columnas del dataframe derecho
    columnas_derecho = com_aux.columns

    # concatenar las columnas del dataframe derecho al inicio del dataframe izquierdo
    pivote_comercial = pd.concat([merge_df[columnas_derecho], merge_df.drop(columnas_derecho, axis=1)], axis=1)
    pivote_comercial['Energetico'].fillna('energia_comercial',inplace=True)
    pivote_comercial['Uso'].fillna('Total',inplace=True)

    pivote_comercial['Uso_num'] = np.where(pivote_comercial['Uso']=='Total', 0, 1)
    pivote_comercial = pivote_comercial.sort_values(by=['Uso_num', 'Uso','Energetico'])
    pivote_comercial.reset_index(inplace=True)
    pivote_comercial.drop(columns=['Variable','index','Nombre_demanda','Uso_num'],inplace=True)
    pivote_comercial = pivote_comercial[['Uso', 'Energetico'] + [col for col in pivote_comercial.columns if col not in ['Energetico', 'Uso']]]    


    sheet_name = 'salida_demanda_comercial'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color

    for j, header in enumerate(pivote_comercial.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_comercial.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    ##Energia residencial
    ###Factor residencial
    salidas_factor_residencial = datos[(datos.Variable == 'calefaccion') | (datos.Variable == 'ACS') | (datos.Variable == 'coccion') | (datos.Variable == 'electricidad_res')]
    salidas_factor_residencial.fillna('',inplace=True)
    salidas_factor_residencial.columns = ['Año','factor','Unidad','Valor']
    salidas_factor_residencial['Valor'] = salidas_factor_residencial['Valor'].replace('', np.nan)
    salidas_factor_residencial['Valor'] = salidas_factor_residencial['Valor'].astype(float)
    salidas_factor_residencial['Valor'] = salidas_factor_residencial['Valor']-1.0

    pivote_factor_residencial = salidas_factor_residencial.pivot_table(values='Valor', index=['factor', 'Unidad'], columns='Año')
    pivote_factor_residencial.reset_index(inplace=True)
    pivote_factor_residencial.rename_axis(None, axis=1, inplace=True)
    pivote_factor_residencial.drop(columns='Unidad',inplace=True)

    sheet_name = 'tasa_crecimiento_residencial'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color
                            
    for j, header in enumerate(pivote_factor_residencial.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_factor_residencial.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    ###demanda residencial
    datos_residencial = pd.read_excel(path,sheet_name='demanda_base_residencial')
    diccionario_residencial = pd.read_excel(path,sheet_name='dic_residencial')
    tasa_residencial = pivote_factor_residencial.copy()
    salidas_residencial = calcular_demanda_residencial(datos_residencial,diccionario_residencial,tasa_residencial)
    regiones_ord = {'RM':0,'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7,'VIII':8,'IX':9,'X':10,'XI':11,'XII':12,'XIV':14,'XV':15,'XVI':16}
    salidas_residencial['orden_reg'] = salidas_residencial['Region'].map(regiones_ord)

    salidas_residencial.sort_values(['orden_reg','Uso'],inplace=True)
    salidas_residencial.reset_index(inplace=True)
    salidas_residencial.drop(columns=['orden_reg','index'],inplace=True)

    sheet_name = 'salida_demanda_residencial'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color
                            
    for j, header in enumerate(salidas_residencial.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in salidas_residencial.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)



    ##Energia publico
    salidas_publico = datos[(datos.Variable == 'energia_publico')]

    pivote_publico = salidas_publico.pivot_table(values='Valor', index=['Variable', 'Unidad'], columns='Año')
    pivote_publico.reset_index(inplace=True)
    pivote_publico.rename_axis(None, axis=1, inplace=True)

    sheet_name = 'salida_demanda_publico'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color
                            
    for j, header in enumerate(pivote_publico.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_publico.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)


    #Transporte
    salidas_transporte = datos[(datos['Variable'].str.startswith('km_')) | (datos['Variable'].str.startswith('pkm_') )| (datos['Variable'].str.startswith('tkm_'))]

    pivote_transporte = salidas_transporte.pivot_table(values='Valor', index=['Variable', 'Unidad'], columns='Año')
    pivote_transporte.reset_index(inplace=True)
    pivote_transporte.rename_axis(None, axis=1, inplace=True)

    sheet_name = 'salida_proyeccion_transporte'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color
                            
    for j, header in enumerate(pivote_transporte.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_transporte.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    #Sector
    salidas_sector = datos[(datos.Unidad == 'miles ton')|(datos.Variable == 'manufactura')|(datos.Variable == 'cemento')|(datos.Variable == 'hierro')|(datos.Variable == 'mineral')|(datos.Variable == 'manufactura_existente')|(datos.Variable == 'manufactura_nuevo')|(datos.Variable == 'cemento_existente')|(datos.Variable == 'cemento_nuevo')|(datos.Variable == 'hierro_existente')|(datos.Variable == 'hierro_nuevo')|(datos.Variable == 'mineral_existente')|(datos.Variable == 'mineral_nuevo')]

    pivote_sector = salidas_sector.pivot_table(values='Valor', index=['Variable', 'Unidad'], columns='Año')
    pivote_sector.reset_index(inplace=True)
    pivote_sector.rename_axis(None, axis=1, inplace=True)

    sheet_name = 'salida_proyeccion_sector'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color

    for j, header in enumerate(pivote_sector.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_sector.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)


    #Cobre
    ##Datos
    participacion_cobre = pd.read_excel(path,sheet_name='participacion_cobre')
    leyes_cobre = pd.read_excel(path,sheet_name='leyes_cobre')
    recuperacion_cobre = pd.read_excel(path,sheet_name='recuperacion_cobre')

    ##Produccion cobre fino
    #datos se llama la variable con datos_proyeccion_demand
    salidas_cobre = datos[datos.Variable == 'cobre']
    salidas_cobre = salidas_cobre.rename(columns={'Valor': 'cobre_total'})

    Rajo = cambiar_estructura(participacion_cobre.copy(),'par_Rajo').reset_index()
    Rajo = Rajo[Rajo['Proceso']=='Mina Rajo'].drop(columns='Proceso')

    Subterranea = cambiar_estructura(participacion_cobre.copy(),'par_Subterranea').reset_index()
    Subterranea = Subterranea[Subterranea['Proceso']=='Mina Subterranea'].drop(columns='Proceso')     

    oxidos = cambiar_estructura(participacion_cobre.copy(),'par_oxidos').reset_index()
    oxidos = oxidos[oxidos['Proceso']=='Óxidos'].drop(columns='Proceso')

    sulfuro = cambiar_estructura(participacion_cobre.copy(),'par_sulfuro').reset_index()
    sulfuro = sulfuro[sulfuro['Proceso']=='Sulfuros'].drop(columns='Proceso')

    refineria = cambiar_estructura(participacion_cobre.copy(),'par_refineria').reset_index()
    refineria = refineria[refineria['Proceso']=='Refinería "Saturación"'].drop(columns='Proceso')

    fundicion = cambiar_estructura(participacion_cobre.copy(),'par_fundicion').reset_index()
    fundicion = fundicion[fundicion['Proceso']=='Fundición "Saturación"'].drop(columns='Proceso')

    ley_oxidos = cambiar_estructura(leyes_cobre.copy(),'ley_oxidos').reset_index()
    ley_oxidos = ley_oxidos[ley_oxidos['Ley']=='Óxidos (LX)'].drop(columns='Ley')

    ley_sulfuros = cambiar_estructura(leyes_cobre.copy(),'ley_sulfuros').reset_index()
    ley_sulfuros = ley_sulfuros[ley_sulfuros['Ley']=='Sulfuros (Concentrado)'].drop(columns='Ley')

    recuperacion_oxidos = cambiar_estructura(recuperacion_cobre.copy(),'recuperacion_oxidos').reset_index()
    recuperacion_oxidos = recuperacion_oxidos[recuperacion_oxidos['Proceso']=='Óxidos'].drop(columns='Proceso')

    recuperacion_sulfuros = cambiar_estructura(recuperacion_cobre.copy(),'recuperacion_sulfuros').reset_index()
    recuperacion_sulfuros = recuperacion_sulfuros[recuperacion_sulfuros['Proceso']=='Sulfuros'].drop(columns='Proceso')

    demanda_cobre = salidas_cobre.merge(oxidos,how='outer')
    demanda_cobre = demanda_cobre.merge(Rajo,how='outer')
    demanda_cobre = demanda_cobre.merge(Subterranea,how='outer')
    demanda_cobre = demanda_cobre.merge(sulfuro,how='outer')
    demanda_cobre = demanda_cobre.merge(refineria,how='outer')
    demanda_cobre = demanda_cobre.merge(fundicion,how='outer')
    demanda_cobre = demanda_cobre.merge(ley_oxidos,how='outer')
    demanda_cobre = demanda_cobre.merge(ley_sulfuros,how='outer')
    demanda_cobre = demanda_cobre.merge(recuperacion_oxidos,how='outer')
    demanda_cobre = demanda_cobre.merge(recuperacion_sulfuros,how='outer')

    prod_total = demanda_cobre['cobre_total']
    part_Rajo = demanda_cobre['par_Rajo']
    part_Subterranea = demanda_cobre['par_Subterranea']
    part_oxidos = demanda_cobre['par_oxidos']
    part_sulfuro = demanda_cobre['par_sulfuro']
    part_refineria = demanda_cobre['par_refineria']
    part_fundicion = demanda_cobre['par_fundicion']
    ley_oxid = demanda_cobre['ley_oxidos']
    ley_sulf = demanda_cobre['ley_sulfuros']
    rec_oxidos = demanda_cobre['recuperacion_oxidos']
    rec_sulfuros = demanda_cobre['recuperacion_sulfuros']

    #Produccion
    demanda_cobre['Óxidos'] = prod_total * part_oxidos
    demanda_cobre['Refinería'] = prod_total * part_refineria
    demanda_cobre['Fundición'] = prod_total * part_fundicion
    demanda_cobre['Unidad'] = 'Miles ton Cu fino'
    demanda_cobre = demanda_cobre[['Año','Unidad','Óxidos','Refinería','Fundición']]

    df_oxidos = demanda_cobre[['Año','Unidad','Óxidos']].rename(columns={'Óxidos':'Valor'})
    df_oxidos['Proceso'] = 'Óxidos'
    df_refineria = demanda_cobre[['Año','Unidad','Refinería']].rename(columns={'Refinería':'Valor'})
    df_refineria['Proceso'] = 'Refinería'
    df_fundicion = demanda_cobre[['Año','Unidad','Fundición']].rename(columns={'Fundición':'Valor'})
    df_fundicion['Proceso'] = 'Fundición'

    df_produccion = df_oxidos.merge(df_refineria,how='outer')
    df_produccion = df_produccion.merge(df_fundicion,how='outer')
    pivote_produccion = df_produccion.pivot_table(values='Valor', index=['Proceso', 'Unidad'], columns='Año').reset_index()
    pivote_produccion.rename_axis(None, axis=1, inplace=True)

    sheet_name = 'salida_produccion_cobre'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color

    for j, header in enumerate(pivote_produccion.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_produccion.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    #Material Procesado
    material_procesado = demanda_cobre.copy()
    material_procesado['Óxidos']= prod_total*part_oxidos/(ley_oxid*rec_oxidos)
    material_procesado['Sulfuro']= prod_total*part_sulfuro/(ley_sulf*rec_sulfuros)
    material_procesado['Total'] = material_procesado['Óxidos'] + material_procesado['Sulfuro']
    material_procesado['Mina Rajo'] = part_Rajo * material_procesado['Total']
    material_procesado['Mina Subterranea'] = part_Subterranea * material_procesado['Total']
    material_procesado['Unidad'] = 'Miles ton de roca'
    material_procesado = material_procesado[['Año','Unidad','Óxidos','Sulfuro','Mina Rajo','Mina Subterranea','Total']]

    df_oxid = material_procesado[['Año','Unidad','Óxidos']].rename(columns={'Óxidos':'Valor'})
    df_oxid['Proceso'] = 'Óxidos'
    df_sulf = material_procesado[['Año','Unidad','Sulfuro']].rename(columns={'Sulfuro':'Valor'})
    df_sulf['Proceso'] = 'Sulfuros'
    df_Rajo = material_procesado[['Año','Unidad','Mina Rajo']].rename(columns={'Mina Rajo':'Valor'})
    df_Rajo['Proceso'] = 'Mina Rajo'
    df_Subterranea = material_procesado[['Año','Unidad','Mina Subterranea']].rename(columns={'Mina Subterranea':'Valor'})
    df_Subterranea['Proceso'] = 'Mina Subterranea'
    df_tot = material_procesado[['Año','Unidad','Total']].rename(columns={'Total':'Valor'})
    df_tot['Proceso'] = 'Total'

    df_mat_procesado = df_oxid.merge(df_sulf,how='outer')
    df_mat_procesado = df_mat_procesado.merge(df_Rajo,how='outer')
    df_mat_procesado = df_mat_procesado.merge(df_Subterranea,how='outer')
    df_mat_procesado = df_mat_procesado.merge(df_tot,how='outer')
    pivote_procesado = df_mat_procesado.pivot_table(values='Valor', index=['Proceso', 'Unidad'], columns='Año').reset_index()
    pivote_procesado.rename_axis(None, axis=1, inplace=True)
    ord_procesado = {'Óxidos':1,'Sulfuros':2,'Mina Rajo':3,'Mina Subterranea':4,'Total':5}
    pivote_procesado['orden'] = pivote_procesado["Proceso"].map(ord_procesado)
    pivote_procesado.sort_values('orden',inplace=True)
    pivote_procesado.reset_index(inplace=True)
    pivote_procesado.drop(columns=['orden','index'],inplace=True)

    sheet_name = 'salida_material_procesado'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    sheet.sheet_properties.tabColor = color

    for j, header in enumerate(pivote_procesado.columns, start=1):
        sheet.cell(row=1, column=j, value=header)
        sheet.cell(row=1, column=j).fill = light_fill
    for i, row in pivote_procesado.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)


    wb.save(path)
    wb.close()



#Importamos los datos historicos y del modelo PMR para proyectar demandas
def importar_datos(path,obs):
    
    #separador = os.path.sep
    #dir_actual = os.path.dirname(os.path.abspath('proyecciones_demanda.py'))
    #dir = separador.join(dir_actual.split(separador)[:-1])
    #path = dir+obs
    
    data = transformar_entrada(path,obs)
    
    param = pd.read_excel(obs,sheet_name = 'Parametros_aux',usecols=[0,1])
    param.columns= ['Parametro','Valor']
    param.dropna(inplace=True)
    param.set_index('Parametro',inplace=True)
    
    param_e = pd.read_excel(obs, sheet_name = 'Parametros_aux',usecols = [5,6])
    param_e.dropna(inplace=True)
    param_e.columns= ['Parametro','Valor']
    param_e.set_index('Parametro',inplace=True)
    
    Panel_medidas = pd.read_excel(obs,sheet_name = 'Panel Medidas',usecols=[0,1],skipfooter=36)
    Panel_medidas.set_index('Nombre',inplace=True)
    
    Edif_sost = pd.read_excel(obs,sheet_name = '4.-Edificacion_sostenible',usecols=[21,57,59,60,61],skiprows=5,nrows=34)
    Edif_sost.columns = ['Año','ahorro consumo CE', 'ahorro consumo RT','ahorro Calef distrital', 'ahorro geotermia']
    Edif_sost.set_index('Año',inplace=True)


    #cols = [i for i in range(5,23)]
    #cols.insert(0,1)
    #demanda_com = pd.read_excel(path, sheet_name= 'DEMANDA_COM_AUX',usecols=cols,nrows=35)
    #demanda_com.set_index('Año',inplace=True)


    param_cemento = pd.read_excel(path,sheet_name= 'parametros_cemento')
    param_cemento.columns = ['Parametro','Valor']
    param_cemento.set_index('Parametro',inplace=True)

    param_mineral = pd.read_excel(path,sheet_name= 'parametros_minas_varias')
    param_mineral.columns = ['Parametro','Valor']
    param_mineral.set_index('Parametro',inplace=True)

    param_manufactura = pd.read_excel(path,sheet_name='parametros_industrias_varias')
    param_manufactura.columns = ['Parametro','Valor']
    param_manufactura.set_index('Parametro',inplace=True)

    dic_comercial = pd.read_excel(path, sheet_name='dic_comercial')
    dic_comercial.set_index('Nombre_demanda',inplace=True)
    
    eficiencia_com = pd.read_excel(path, sheet_name='eficiencia_uso_comercial')
    participacion_uso = pd.read_excel(path, sheet_name='participacion_uso_comercial')
    participacion_energetico = pd.read_excel(path, sheet_name='participacion_energeticos_com')

    datos = {'data':data, 'param': param, 'Panel_medidas': Panel_medidas, 'Edif_sost' : Edif_sost, 'param_e': param_e, 'Cemento':param_cemento,'Mineral':param_mineral,'Manufactura':param_manufactura,
    'dic_comercial':dic_comercial,'eficiencia_com': eficiencia_com, 'participacion_uso': participacion_uso, 'participacion_energetico': participacion_energetico}
    return datos

#Armamos un dataframe con las proyecciones de demanda transporte a partir del modelo STEP3
def proyecciones_step(path_fold = r'datos_proyecciones/',path_pkm = r'PKM_Tecnologia_todos.csv', path_tkm = r'TKM_Tecnologia_todos.csv', path_vkm = r'VKM_Tecnologia_todos.csv',path=None):

    pkm_df = pd.read_csv(path_fold+path_pkm)
    tkm_df = pd.read_csv(path_fold+path_tkm)
    vkm_df = pd.read_csv(path_fold+path_vkm)
    transantiago_aux_df = pd.read_excel(path, sheet_name='aux_transantiago')
    
    '''Condiciones para filtrar dataframes'''
    años_cond_pkm = (pkm_df['Año'] <= 2050) &(pkm_df['Año'] >= 2017)
    urbano_pkm = (pkm_df['Ambito'] == 'urbano')
    interurbano_pkm = (pkm_df['Ambito'] == 'interurbano')

    años_cond_tkm = (tkm_df['Año'] <= 2050) &(tkm_df['Año'] >= 2017)
    urbano_tkm = (tkm_df['Ambito'] == 'urbano')
    interurbano_tkm = (tkm_df['Ambito'] == 'interurbano')

    años_cond_vkm = (vkm_df['Año'] <= 2050) &(vkm_df['Año'] >= 2017)
    interurbano_vkm = (vkm_df['Ambito'] == 'interurbano')
    urbano_vkm = (vkm_df['Ambito'] == 'urbano')
    regiones = (vkm_df['Region'] != 'RM')
    RM = (vkm_df['Region'] == 'RM')

    ferroviario_pkm =  (pkm_df['Modo'] == 'Tren_urbano')
    ferroviario_tkm = (tkm_df['Modo'] == 'Ferroviario')
    maritimo = (tkm_df['Tipo'] == 'Carga-Maritimo')
    petroleo_comb = (tkm_df['Motorizacion'] == 'Petróleo Combustible')
    Aereo = (pkm_df['Tipo']=='Pasajero-Aereo' )
    no_internacional = (pkm_df['Ambito']!= 'internacional')
    #no_GRL = (tkm_df['Modo']!= 'GRL')


    moto =  (vkm_df['Modo'] == 'Motocicleta')
    tractocamion = (vkm_df['Modo'] == 'Camión pesado')
    camion = (vkm_df['Modo'] == 'Camión liviano') | (vkm_df['Modo'] == 'Camión mediano') #Se debe incluir medianos?
    bus = (vkm_df['Modo'] == 'Bus Rígido')
    transantiago = (vkm_df['Modo'] == 'Bus Articulado')
    taxi = (vkm_df['Modo'] == 'Taxi')
    vehiculos = (vkm_df['Modo'] == 'Vehículo Liviano Pasajero')
    medianos = (vkm_df['Modo'] == 'Vehículo Liviano Comercial')

    col_pkm = ['Año','Modo','Ambito','PKM_Tecnologia']
    col_vkm = ['Año','Modo','Ambito','VKM_Tecnologia']
    col_aereo =['Tipo','Año','Modo','Ambito','PKM_Tecnologia']
    col_tkm = ['Año','Modo','Ambito','TKM_Tecnologia']
    col_mar = ['Tipo','Año','Modo','Ambito','TKM_Tecnologia']
    
    '''Calculamos los pkm, tkm y km (vkm)'''
    tren_urb = pkm_df[col_pkm][años_cond_pkm & urbano_pkm & ferroviario_pkm]
    tren_inter = pkm_df[col_pkm][años_cond_pkm & interurbano_pkm & ferroviario_pkm]
    ferro_tkm = tkm_df[col_tkm][años_cond_tkm & no_internacional & ferroviario_tkm]

    aereo_pkm = pkm_df[col_aereo][años_cond_pkm & no_internacional & Aereo]

    mar_tkm = tkm_df[col_mar][años_cond_tkm & interurbano_tkm & maritimo & petroleo_comb]

    moto_vkm = vkm_df[col_vkm][años_cond_vkm & moto]

    tracto_vkm = vkm_df[col_vkm][años_cond_vkm & tractocamion]
    camion_vkm = vkm_df[col_vkm][años_cond_vkm & camion]

    bus_inter_reg = vkm_df[col_vkm][años_cond_vkm & bus & interurbano_vkm & regiones]
    bus_urb_reg = vkm_df[col_vkm][años_cond_vkm & bus & urbano_vkm & regiones]
    bus_inter_rm = vkm_df[col_vkm][años_cond_vkm & bus & interurbano_vkm & RM]
    bus_urb_rm = vkm_df[col_vkm][años_cond_vkm & bus & urbano_vkm & RM]


    transantiago_vkm = vkm_df[col_vkm][años_cond_vkm & transantiago]
    taxi_vkm = vkm_df[col_vkm][años_cond_vkm & taxi]
    veh_vkm = vkm_df[col_vkm][años_cond_vkm & vehiculos]
    med_vkm = vkm_df[col_vkm][años_cond_vkm & medianos]

    dfs = [[veh_vkm,'km_livianos'],[taxi_vkm,'km_taxi'],[transantiago_vkm,'km_transantiago'],[bus_urb_rm, 'km_bus_urbano_RM'],[bus_inter_rm, 'km_bus_interurbano_RM'],[bus_urb_reg, 'km_bus_urbano_regiones'],[bus_inter_reg, 'km_bus_interurbano_regiones'],[med_vkm,'km_medianos'],[camion_vkm,'km_camion'],[tracto_vkm,'km_tractocamion'],[moto_vkm,'km_moto'],[aereo_pkm, 'pkm_aereo'],[mar_tkm,'tkm_maritimo'],[tren_urb,'pkm_ferroviario_urb'],[tren_inter,'pkm_ferroviario_inter'],[ferro_tkm, 'tkm_ferroviario']]

    df_list = []
    for sets in dfs:
        if sets[1] == 'pkm_aereo' or sets[1] == 'tkm_maritimo':
            sets[0].drop(columns=['Tipo'],inplace=True)
        sets[0].drop(columns=['Modo','Ambito'],inplace=True)
        sets[0] = sets[0].groupby('Año').sum()
        sets[0].reset_index(inplace=True)
        sets[0].columns = ['Año',sets[1]]
        sets[0][sets[1]] = sets[0][sets[1]]/1000
        df_list.append(sets[0])
    
    '''Unimos los dataframes'''
    df_merged = reduce(lambda  left,right: pd.merge(left,right,on=['Año'],how='outer'), df_list)
    df_merged = df_merged.astype(int)
    df_merged.set_index('Año',inplace=True)
    #Separar miles utilzando ","
    #df_merged.loc[:,df_merged.columns!="Año"] = df_merged.loc[:,df_merged.columns!="Año"].applymap('{:,}'.format)

    #Ajustamos transantiago y buses rigidos rm urbano
    df_merged.reset_index(inplace=True)
    df_merged = df_merged.merge(transantiago_aux_df,how='left')

    df_merged['km_transantiago'] = df_merged['km_transantiago'] + df_merged['km_bus_urbano_RM']* df_merged['Fraccion Rigidos']
    df_merged['km_bus_urbano_RM'] = df_merged['km_bus_urbano_RM']*(1-df_merged['Fraccion Rigidos'])
    df_merged.drop(columns='Fraccion Rigidos',inplace=True)
    df_merged.set_index('Año',inplace=True)

    return df_merged



#Funciones para calcular proyecciones de demanda
def calcular_PIB(data,año):
    tasa_pib = data.loc[año]['tasa PIB']
    pib_anterior = data.loc[año-1]['PIB']
    pib = pib_anterior * (1+ tasa_pib )
    return pib

def calcular_PIB_millUSD(data,param,año):
    pib = data.loc[año]['PIB']
    precio_usd = data.loc[año]['Tipo cambio dólar']
    pib_usd = (pib/precio_usd)*1000
    return pib_usd

def calcular_HAB_viv(data,param,año):
    pob = data.loc[año]['Población']
    viv = data.loc[año]['Viviendas']
    if pd.isna(viv):
        elast = param.loc['elasticidad ']['Valor']
        tasa_pib = data.loc[año]['tasa PIB']
        hab_viv_anterior = data.loc[año-1]['HAB/viv']
        Hab_viv = (1+elast*tasa_pib)*hab_viv_anterior
    if pd.notna(viv):
        Hab_viv = pob/viv
    return Hab_viv

def calcular_viviendas(data,año):
    pob = data.loc[año]['Población']
    hab_viv = data.loc[año]['HAB/viv']
    viv = round(pob/hab_viv,0)
    return viv

def calcular_pib_pc(data,año):
    pob = data.loc[año]['Población']
    pib_usd = data.loc[año]['PIB (mill USD)']
    pib_pc = round((pib_usd/pob)*1000000,0)
    return pib_pc



def calcular_tasa_crecimiento_calefaccion_residencial(data,Panel,Edif,año,año_hist):
    año_hist = list(data.index.values)[0]
    viv_hist = data.loc[año_hist]['Viviendas']
    viv = data.loc[año]['Viviendas']

    fom_renovacion_e = Panel.loc['Fomento a la renovación energética de viviendas']['Activar'] == 'si'
    reac_ter_viv = Panel.loc['Reacondicionamiento térmico viviendas vulnerables']['Activar'] == 'si'
    calef_dist = Panel.loc['Calefacción Distrital']['Activar'] == 'si'
    geotermia = Panel.loc['Geotermia']['Activar'] == 'si'
    
    ahorro_consumo_ce = Edif.loc[año]['ahorro consumo CE'] if fom_renovacion_e else 0
    ahorro_consumo_rt = Edif.loc[año]['ahorro consumo RT'] if reac_ter_viv else 0
    ahorro_calef_dist = Edif.loc[año]['ahorro Calef distrital'] if calef_dist else 0
    ahorro_geotermia = Edif.loc[año]['ahorro geotermia'] if geotermia else 0
    
    frac_viv  = (viv/viv_hist)
    calefaccion = frac_viv * (1 - ahorro_consumo_ce) * (1 - ahorro_consumo_rt) * (1 - ahorro_calef_dist) * (1 - ahorro_geotermia)
    tasa_crecimiento_calefaccion = calefaccion -1
    return tasa_crecimiento_calefaccion



def calcular_tasa_crecimiento_ACS_residencial(data,año,año_hist):
    año_hist = list(data.index.values)[0]
    viv_hist = data.loc[año_hist]['Viviendas']
    viv = data.loc[año]['Viviendas']
    
    tasa_crecimiento_ACS = viv/viv_hist -1
    return tasa_crecimiento_ACS

def calcular_tasa_crecimiento_coccion_residencial(data,año,año_hist):
    año_hist = list(data.index.values)[0]
    viv_hist = data.loc[año_hist]['Viviendas']
    viv = data.loc[año]['Viviendas']
    
    tasa_crecimiento_coccion = viv/viv_hist -1
    return tasa_crecimiento_coccion


def calcular_demanda_residencial(datos_residencia,diccionario_residencial, factor_residencial):
    #Reestructuramos los datos
    diccionario_residencial_df = diccionario_residencial.set_index(['Region','Uso','Energetico'])
    datos_res = cambiar_estructura(datos_residencia,'demanda_base')
    resultado = datos_res.join(diccionario_residencial_df, how='left')
    factor_residencial = cambiar_estructura(factor_residencial,'tasa_crecimiento')
    resultado = resultado.join(factor_residencial,how='left',on=['factor','Año'])

    #Comenzamos el cálculo
    demanda_base = resultado['demanda_base']
    tasa_crecimiento = resultado['tasa_crecimiento']

    resultado['demanda_util'] = demanda_base * (1 + tasa_crecimiento)
    
    #Reestructuramos la salida
    resultado.reset_index(inplace=True)
    resultado.drop(columns=['factor','demanda_base','tasa_crecimiento'],inplace=True)
    resultado['Unidad'] = 'Tcal'
    resultado = resultado.pivot_table(values='demanda_util', index=['Region','Uso', 'Energetico','Unidad'], columns='Año')
    resultado.reset_index(inplace=True)
    resultado.rename_axis(None, axis=1, inplace=True)
    return resultado  


def calcular_energia_comercial(data,param_e,año):
    alfa = param_e.loc['alfa']['Valor']
    beta = param_e.loc['beta']['Valor']
    pib= data.loc[año]['PIB']
    
    energia_comercial = np.exp(alfa+beta*np.log(pib))
    return energia_comercial




def calcular_cemento(data,param_sector,año,año_hist):
    
    #Traemos los parámetros del modelo econométrico
    param_cemento = param_sector
    
    a= param_cemento.loc['a']['Valor']
    b= param_cemento.loc['b']['Valor']
    c= param_cemento.loc['c']['Valor']
    d= param_cemento.loc['d']['Valor']
    
    
    pib = data.loc[año]['PIB']
    pib_anterior = data.loc[año-1]['PIB']
    cemento_anterior = data.loc[año-1]['cemento']

    cemento = np.exp(a + np.log(cemento_anterior)*b + np.log(pib)*c + np.log(pib_anterior)*d)
    return cemento

def calcular_mineral(data,param_sector,año):

    #Traemos los parámetros del modelo econométrico
    param_mineral = param_sector

    a= param_mineral.loc['a']['Valor']
    b= param_mineral.loc['b']['Valor']
    pib = data.loc[año]['PIB']
    salitre = data.loc[año]['salitre']

    mineral = np.exp(np.log(pib)*b+a) - salitre
    return mineral


def calcular_manufactura(data,param_sector,año):
    
    #Traemos los parámetros del modelo econométrico
    param_manufactura = param_sector
    
    a= param_manufactura.loc['a']['Valor']
    b= param_manufactura.loc['b']['Valor']
    c= param_manufactura.loc['c']['Valor']
    
    
    pib = data.loc[año]['PIB']
    manufactura_anterior = data.loc[año-1]['manufactura']    

    manufactura = np.exp(np.log(pib)*c + b*np.log(manufactura_anterior) + a)
    return manufactura


def calcular_energia_publico(data,año,año_hist):
    pib_hist= data.loc[año_hist]['PIB']
    pib= data.loc[año]['PIB']
    energ_pub_hist = data.loc[año_hist]['energia_publico']
    
    energia_publico = (pib/pib_hist)*energ_pub_hist
    return energia_publico


def calcular_manufactura_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    manufactura = data.loc[año]['manufactura']
    manufactura_hist = data.loc[año_hist_data]['manufactura']
    if año <= año_hist:
        manufactura_v = data.loc[año]['manufactura']
    else:
        if manufactura > manufactura_hist:
            manufactura_v = manufactura_hist
        else:
            manufactura_v = manufactura
    return manufactura_v

def calcular_celulosa_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    celulosa = data.loc[año]['celulosa']
    celulosa_hist = data.loc[año_hist_data]['celulosa']
    if año <= año_hist:
        celulosa_v = data.loc[año]['celulosa']
    else:
        if celulosa > celulosa_hist:
            celulosa_v = celulosa_hist
        else:
            celulosa_v = celulosa
    return celulosa_v

def calcular_azucar_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    azucar = data.loc[año]['azucar']
    azucar_hist = data.loc[año_hist_data]['azucar']
    if año <= año_hist:
        azucar_v = data.loc[año]['azucar']
    else:
        if azucar > azucar_hist:
            azucar_v = azucar_hist
        else:
            azucar_v = azucar
    return azucar_v

def calcular_cemento_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    cemento = data.loc[año]['cemento']
    cemento_hist = data.loc[año_hist_data]['cemento']
    if año <= año_hist:
        cemento_v = data.loc[año]['cemento']
    else:
        if cemento > cemento_hist:
            cemento_v = cemento_hist
        else:
            cemento_v = cemento
    return cemento_v
    
def calcular_hierro_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    hierro = data.loc[año]['hierro']
    hierro_hist = data.loc[año_hist_data]['hierro']
    if año <= año_hist:
        hierro_v = data.loc[año]['hierro']
    else:
        if hierro > hierro_hist:
            hierro_v = hierro_hist
        else:
            hierro_v = hierro
    return hierro_v

def calcular_pesca_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    pesca = data.loc[año]['pesca']
    pesca_hist = data.loc[año_hist_data]['pesca']
    if año <= año_hist:
        pesca_v = data.loc[año]['pesca']
    else:
        if pesca > pesca_hist:
            pesca_v = pesca_hist
        else:
            pesca_v = pesca
    return pesca_v

def calcular_acero_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    acero = data.loc[año]['acero']
    acero_hist = data.loc[año_hist_data]['acero']
    if año <= año_hist:
        acero_v = data.loc[año]['acero']
    else:
        if acero > acero_hist:
            acero_v = acero_hist
        else:
            acero_v = acero
    return acero_v

def calcular_petroquimica_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    petroquimica = data.loc[año]['petroquimica']
    petroquimica_hist = data.loc[año_hist_data]['petroquimica']
    if año <= año_hist:
        petroquimica_v = data.loc[año]['petroquimica']
    else:
        if petroquimica > petroquimica_hist:
            petroquimica_v = petroquimica_hist
        else:
            petroquimica_v = petroquimica
    return petroquimica_v

def calcular_mineral_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    mineral = data.loc[año]['mineral']
    mineral_hist = data.loc[año_hist_data]['mineral']
    if año <= año_hist:
        mineral_v = data.loc[año]['mineral']
    else:
        if mineral > mineral_hist:
            mineral_v = mineral_hist
        else:
            mineral_v = mineral
    return mineral_v

def calcular_cobre_existente(data,año,año_hist):
    año_hist_data=  list(data.index.values)[0]
    cobre = data.loc[año]['cobre']
    cobre_hist = data.loc[año_hist_data]['cobre']
    if año <= año_hist:
        cobre_v = data.loc[año]['cobre']
    else:
        if cobre > cobre_hist:
            cobre_v = cobre_hist
        else:
            cobre_v = cobre
    return cobre_v


def calcular_manufactura_nuevo(data,año):
    manufactura = data.loc[año]['manufactura']
    manufactura_exist = data.loc[año]['manufactura_existente']
    
    manufactura_nuevo = manufactura-manufactura_exist
    return manufactura_nuevo

def calcular_celulosa_nuevo(data,año):
    celulosa = data.loc[año]['celulosa']
    celulosa_exist = data.loc[año]['celulosa_existente']
    
    celulosa_nuevo = celulosa-celulosa_exist
    return celulosa_nuevo

def calcular_azucar_nuevo(data,año):
    azucar = data.loc[año]['azucar']
    azucar_exist = data.loc[año]['azucar_existente']
    
    azucar_nuevo = azucar-azucar_exist
    return azucar_nuevo

def calcular_cemento_nuevo(data,año):
    cemento = data.loc[año]['cemento']
    cemento_exist = data.loc[año]['cemento_existente']
    
    cemento_nuevo = cemento-cemento_exist
    return cemento_nuevo

def calcular_hierro_nuevo(data,año):
    hierro = data.loc[año]['hierro']
    hierro_exist = data.loc[año]['hierro_existente']
    
    hierro_nuevo = hierro-hierro_exist
    return hierro_nuevo

def calcular_pesca_nuevo(data,año):
    pesca = data.loc[año]['pesca']
    pesca_exist = data.loc[año]['pesca_existente']
    
    pesca_nuevo = pesca-pesca_exist
    return pesca_nuevo

def calcular_acero_nuevo(data,año):
    acero = data.loc[año]['acero']
    acero_exist = data.loc[año]['acero_existente']
    
    acero_nuevo = acero-acero_exist
    return acero_nuevo

def calcular_petroquimica_nuevo(data,año):
    petroquimica = data.loc[año]['petroquimica']
    petroquimica_exist = data.loc[año]['petroquimica_existente']
    
    petroquimica_nuevo = petroquimica-petroquimica_exist
    return petroquimica_nuevo

def calcular_mineral_nuevo(data,año):
    mineral = data.loc[año]['mineral']
    mineral_exist = data.loc[año]['mineral_existente']
    
    mineral_nuevo = mineral-mineral_exist
    return mineral_nuevo

def calcular_cobre_nuevo(data,año):
    cobre = data.loc[año]['cobre']
    cobre_exist = data.loc[año]['cobre_existente']
    
    cobre_nuevo = cobre-cobre_exist
    return cobre_nuevo


def calcular_demanda_comercial(energia_util,dic_comercial,eficiencia_com,participacion_uso,participacion_energetico):
    for i in dic_comercial.index:
        for año in energia_util.index:
            
            eficiencia = float(eficiencia_com[(eficiencia_com['Energetico'] == dic_comercial.loc[i]['Energetico']) & (eficiencia_com['Uso'] == dic_comercial.loc[i]['Uso'])][año])
            part_energetico = float(participacion_energetico[(participacion_energetico['Energetico'] == dic_comercial.loc[i]['Energetico'])][año])
            part_uso = float(participacion_uso[(participacion_uso['Energetico'] == dic_comercial.loc[i]['Energetico']) & (participacion_uso['Uso'] == dic_comercial.loc[i]['Uso'])][año])
            energia_final_total = energia_util.loc[año]['energia_comercial']
            #Energia util por uso final y energético
            energia_util.loc[año,i] = energia_final_total * part_uso* part_energetico * eficiencia


def obtener_año_hist(path,hoja_excel):
    df = pd.read_excel(path, sheet_name=hoja_excel)

    var_eliminar = ['Unidad','Fuente']
    data = df.copy()
    for v in var_eliminar:
        if v in df.columns:
            data.drop(columns=v,inplace=True)

    # Obtener el año histórico para cada variable  # Primera columna con los nombres de las variables
    #datos = df.iloc[:, 3:]  # Columnas a partir de la cuarta contienen los datos
    #columnas= [x for x in datos.columns]
    #columnas.insert(0,'Variable')
    data = data.rename(columns={'Variable':'Año'})

    df_historico = data.transpose().reset_index()
    df_historico.columns = df_historico.iloc[0]
    df_historico = df_historico.iloc[1:]

    # Reiniciar los índices si es necesario
    df_historico = df_historico.reset_index(drop=True)
    df_historico = df_historico.dropna(axis=1, how='all')
    df_historico = df_historico.set_index('Año')

    año_historico = {}

    for columna in df_historico.columns:
        ultimo_indice = df_historico[columna].last_valid_index()
        if isinstance(ultimo_indice, str):  # Si el índice es de tipo string, convertirlo a entero
            ultimo_indice = int(ultimo_indice)
        año_historico[columna] = ultimo_indice

    return df_historico, año_historico

def obtener_año_hist_reg(path,hoja_excel):
    df = pd.read_excel(path, sheet_name=hoja_excel)

    var_eliminar = ['Unidad','Fuente']
    data = df.copy()
    for v in var_eliminar:
        if v in df.columns:
            data.drop(columns=v,inplace=True)

    # Obtener el año histórico para cada variable  # Primera columna con los nombres de las variables
    regiones = ['R1','R2','R3','R4','R5','R6','R7','R8','R9','R10','R11','R12','RM','R14','R15','R16']

    regiones_dicc = {'R1':'I', 'R2':'II','R3':'III','R4':'IV','R5':'V','R6':'VI','R7':'VII',
                        'R8':'VIII','R9':'IX','R10':'X','R11':'XI','R12':'XII','RM':'RM','R14':'XIV','R15':'XV','R16':'XVI'}

    data = data.rename(columns={'Variable':'Año'})
    df_historico_regional = pd.DataFrame()
    año_historico = {}
    for reg in regiones:
        data_reg = data.query('Region == @reg')
        data_reg = data_reg.drop(columns='Region')
        df_historico = data_reg.transpose().reset_index()
        df_historico.columns = df_historico.iloc[0]
        df_historico = df_historico.iloc[1:]

        # Reiniciar los índices si es necesario
        df_historico = df_historico.reset_index(drop=True)
        df_historico = df_historico.dropna(axis=1, how='all')
        df_historico = df_historico.set_index('Año')



        for columna in df_historico.columns:
            ultimo_indice = df_historico[columna].last_valid_index()
            if isinstance(ultimo_indice, str):  # Si el índice es de tipo string, convertirlo a entero
                ultimo_indice = int(ultimo_indice)
            año_historico[columna+'_'+regiones_dicc[reg]] = ultimo_indice
        df_historico['Region'] = reg
        df_historico = df_historico[['Region'] + [col for col in df_historico.columns if col != 'Region']]
        df_historico_regional = pd.concat([df_historico_regional,df_historico],axis=0)

    df_historico_regional = df_historico_regional.reset_index()
    df_historico_regional= df_historico_regional.set_index(['Año'])


    df_historico_regional['Region'] = df_historico_regional['Region'].map(regiones_dicc)
    df_pivoted = df_historico_regional.reset_index().pivot_table(index='Año', columns='Region', values=df_historico_regional.columns[1:])

    column_template = '{}_{}'
    # Crea una lista con los nuevos nombres de las columnas
    new_columns = []
    for col, region in df_pivoted.columns:
        new_columns.append(column_template.format(col, region))

    # Renombra las columnas utilizando la lista de nuevos nombres
    df_pivoted.columns = new_columns

    cols_mod = []
    for elemento in df_pivoted.columns:
        if elemento.endswith("_RM_RM"):
            cols_mod.append(elemento[:-3])
        else:
            if elemento == 'km_transantiago_RM':
                cols_mod.append('km_transantiago')
            else:
                cols_mod.append(elemento)
    df_pivoted.columns = cols_mod 

    #ordenar
    col_ord = ['km_livianos','km_taxi','km_transantiago','km_bus_urbano','km_bus_interurbano',
                'km_medianos','km_camion','km_tractocamion','km_moto','pkm_aereo',
                'tkm_maritimo','pkm_ferroviario_urb','pkm_ferroviario_inter','tkm_ferroviario']
    reg_ord = ['XV','I','II','III','IV','V','RM','VI','VII','XVI','VIII','IX',
                    'XIV','X','XI','XII']
    veh_ord = []
    for c in col_ord:
        if c == 'km_transantiago':
            if c in df_pivoted.columns:
                veh_ord.append(c)
        else:
            for r in reg_ord:
                modo = c+'_'+r
                if modo in df_pivoted.columns:
                    veh_ord.append(c+'_'+r)
                    
    df_pivoted = df_pivoted[veh_ord]
    return df_pivoted, año_historico


#Funcion que aplica todos los calculos, juntando los resultados en un solo dataframe
def calcular_proyecciones(path,obs):
    global dic_comercial
    datos = importar_datos(path,obs)
    global data, param, Panel_medidas, Edif_sost, param_e, datos_cemento, datos_mineral, datos_manufactura, dic_comercial,eficiencia_com,participacion_uso,participacion_energetico  
    data, param, Panel_medidas, Edif_sost, param_e, datos_cemento, datos_mineral, datos_manufactura, dic_comercial,eficiencia_com,participacion_uso,participacion_energetico = datos['data'],datos['param'],datos['Panel_medidas'],datos['Edif_sost'], datos['param_e'],datos['Cemento'],datos['Mineral'],datos['Manufactura'], datos['dic_comercial'],datos['eficiencia_com'],datos['participacion_uso'],datos['participacion_energetico']
    data = data.drop(index=data[data.index < 2017].index)
    años= list(data.index.values)
    _,año_hist = obtener_año_hist(path,'datos_entrada_historicos')
    


    for año in años:
        
        #pib
        pib_v = data.loc[año]['PIB']
        if pd.isna(pib_v):
            data.loc[año,'PIB'] = calcular_PIB(data,año)
            
        #pib_usd
        pib_usd_v = data.loc[año]['PIB (mill USD)']
        if pd.isna(pib_usd_v):
            data.loc[año,'PIB (mill USD)'] = calcular_PIB_millUSD(data,param,año)
        
        #HAB_viv
        hab_viv_v = data.loc[año]['HAB/viv']
        if pd.isna(hab_viv_v):
            data.loc[año,'HAB/viv'] = calcular_HAB_viv(data,param,año)
            
        #Viviendas
        viv_v = data.loc[año]['Viviendas']
        if pd.isna(viv_v):
            data.loc[año,'Viviendas'] = calcular_viviendas(data,año)
            
        #PIB per capita
        pib_pc_v = data.loc[año]['PIB per cápita (usd/hab)']
        if pd.isna(pib_pc_v):
            data.loc[año,'PIB per cápita (usd/hab)'] = calcular_pib_pc(data,año)

        #Calefaccion
        calefaccion_v = data.loc[año]['calefaccion']
        if pd.isna(calefaccion_v):
            data.loc[año,'calefaccion'] = calcular_tasa_crecimiento_calefaccion_residencial(data,Panel_medidas,Edif_sost,año,año_hist['electricidad_res']) +1.0
        
        #ACS
        ACS_v = data.loc[año]['ACS']
        if pd.isna(ACS_v):
            data.loc[año,'ACS'] = calcular_tasa_crecimiento_ACS_residencial(data,año,año_hist['electricidad_res']) +1.0
        
        #coccion
        coccion_v = data.loc[año]['coccion']
        if pd.isna(coccion_v):
            data.loc[año,'coccion'] = calcular_tasa_crecimiento_coccion_residencial(data,año,año_hist['electricidad_res']) +1.0
        
        #energia_comercial
        energia_comercial_v = data.loc[año]['energia_comercial']
        if pd.isna(energia_comercial_v):
            data.loc[año,'energia_comercial'] = calcular_energia_comercial(data,param_e,año)
        
        #energia_comercial
        energia_publico_v = data.loc[año]['energia_publico']
        if pd.isna(energia_publico_v):
            data.loc[año,'energia_publico'] = calcular_energia_publico(data,año,2017)
        
        #cemento
        cemento_v = data.loc[año]['cemento']
        if pd.isna(cemento_v):
            data.loc[año,'cemento'] = calcular_cemento(data,datos_cemento,año,año_hist['cemento'])

        #hierro
        #hierro_v = data.loc[año]['hierro']
        #if pd.isna(hierro_v):
        #    data.loc[año,'hierro'] = hierro.loc[año]['Demanda']

        #mineral
        mineral_v = data.loc[año]['mineral']
        if pd.isna(mineral_v):
            data.loc[año,'mineral'] = calcular_mineral(data,datos_mineral,año)
        
        #manufactura
        manufactura_v = data.loc[año]['manufactura']
        if pd.isna(manufactura_v):
            data.loc[año,'manufactura'] = calcular_manufactura(data,datos_manufactura,año)
        
        #manufactura_existente
        manufactura_existente_v = data.loc[año]['manufactura_existente']
        if pd.isna(manufactura_existente_v):
            data.loc[año,'manufactura_existente'] = calcular_manufactura_existente(data,año,año_hist['manufactura'])
        
        #celulosa_existente
        celulosa_existente_v = data.loc[año]['celulosa_existente']
        if pd.isna(celulosa_existente_v):
            data.loc[año,'celulosa_existente'] = calcular_celulosa_existente(data,año,año_hist['celulosa'])
            
        #azucar_existente
        azucar_existente_v = data.loc[año]['azucar_existente']
        if pd.isna(azucar_existente_v):
            data.loc[año,'azucar_existente'] = calcular_azucar_existente(data,año,año_hist['azucar'])
            
        #cemento_existente
        cemento_existente_v = data.loc[año]['cemento_existente']
        if pd.isna(cemento_existente_v):
            data.loc[año,'cemento_existente'] = calcular_cemento_existente(data,año,año_hist['cemento'])

        #hierro_existente
        hierro_existente_v = data.loc[año]['hierro_existente']
        if pd.isna(hierro_existente_v):
            data.loc[año,'hierro_existente'] = calcular_hierro_existente(data,año,año_hist['hierro'])

        #pesca_existente
        pesca_existente_v = data.loc[año]['pesca_existente']
        if pd.isna(pesca_existente_v):
            data.loc[año,'pesca_existente'] = calcular_pesca_existente(data,año,año_hist['pesca'])

        #acero_existente
        acero_existente_v = data.loc[año]['acero_existente']
        if pd.isna(acero_existente_v):
            data.loc[año,'acero_existente'] = calcular_acero_existente(data,año,año_hist['acero'])

        #petroquimica_existente
        petroquimica_existente_v = data.loc[año]['petroquimica_existente']
        if pd.isna(petroquimica_existente_v):
            data.loc[año,'petroquimica_existente'] = calcular_petroquimica_existente(data,año,año_hist['petroquimica'])

        #mineral_existente
        mineral_existente_v = data.loc[año]['mineral_existente']
        if pd.isna(mineral_existente_v):
            data.loc[año,'mineral_existente'] = calcular_mineral_existente(data,año,año_hist['minas_varias'])
        
        #cobre_existente
        cobre_existente_v = data.loc[año]['cobre_existente']
        if pd.isna(cobre_existente_v):
            data.loc[año,'cobre_existente'] = calcular_cobre_existente(data,año,año_hist['cobre'])

        #manufactura_nuevo
        manufactura_nuevo_v = data.loc[año]['manufactura_nuevo']
        if pd.isna(manufactura_nuevo_v):
            data.loc[año,'manufactura_nuevo'] = calcular_manufactura_nuevo(data,año)
        
        #celulosa_nuevo
        celulosa_nuevo_v = data.loc[año]['celulosa_nuevo']
        if pd.isna(celulosa_nuevo_v):
            data.loc[año,'celulosa_nuevo'] = calcular_celulosa_nuevo(data,año)
        
        #azucar_nuevo
        azucar_nuevo_v = data.loc[año]['azucar_nuevo']
        if pd.isna(azucar_nuevo_v):
            data.loc[año,'azucar_nuevo'] = calcular_azucar_nuevo(data,año)
        
        #cemento_nuevo
        cemento_nuevo_v = data.loc[año]['cemento_nuevo']
        if pd.isna(cemento_nuevo_v):
            data.loc[año,'cemento_nuevo'] = calcular_cemento_nuevo(data,año)
        
        #hierro_nuevo
        hierro_nuevo_v = data.loc[año]['hierro_nuevo']
        if pd.isna(hierro_nuevo_v):
            data.loc[año,'hierro_nuevo'] = calcular_hierro_nuevo(data,año)
        
        #pesca_nuevo
        pesca_nuevo_v = data.loc[año]['pesca_nuevo']
        if pd.isna(pesca_nuevo_v):
            data.loc[año,'pesca_nuevo'] = calcular_pesca_nuevo(data,año)
        
        #acero_nuevo
        acero_nuevo_v = data.loc[año]['acero_nuevo']
        if pd.isna(acero_nuevo_v):
            data.loc[año,'acero_nuevo'] = calcular_acero_nuevo(data,año)
        
        #petroquimica_nuevo
        petroquimica_nuevo_v = data.loc[año]['petroquimica_nuevo']
        if pd.isna(petroquimica_nuevo_v):
            data.loc[año,'petroquimica_nuevo'] = calcular_petroquimica_nuevo(data,año)
        
        #mineral_nuevo
        mineral_nuevo_v = data.loc[año]['mineral_nuevo']
        if pd.isna(mineral_nuevo_v):
            data.loc[año,'mineral_nuevo'] = calcular_mineral_nuevo(data,año)
        
        #cobre_nuevo
        cobre_nuevo_v = data.loc[año]['cobre_nuevo']
        if pd.isna(cobre_nuevo_v):
            data.loc[año,'cobre_nuevo'] = calcular_cobre_nuevo(data,año)

    #demanda comercial
    calcular_demanda_comercial(data,dic_comercial,eficiencia_com,participacion_uso,participacion_energetico)
        
    #Calculo de proyecciones transporte nacional
    transporte_nacional = proyecciones_step(path=path) #Este debo modificar
    df_hist_transporte, año_hist_transporte = obtener_año_hist(path,'datos_historicos_transporte')
    cols_historicas = df_hist_transporte.columns
    for col in cols_historicas:
        print(f'Agregando información histórica de {col}')
        for año in list(transporte_nacional.index.values):
            if año <= año_hist_transporte[col]:
                transporte_historico = df_hist_transporte.loc[año][col]
                if not pd.isna(transporte_historico):
                    transporte_nacional.loc[año,col] = transporte_historico




    #Eliminamos columna auxiliar salitre
    data = data.drop(columns='salitre')

    #Armando el dataframe
    data_proyecciones_demanda = data.reset_index().merge(transporte_nacional.reset_index(),how='inner')

    return data_proyecciones_demanda



def split_column(col):
    split = re.split(r'_(\w+)$', col)
    return split[0],split[1]



def proyecciones_step_regiones(path_fold = r'datos_proyecciones/',path_pkm = r'PKM_Tecnologia_todos.csv', path_tkm = r'TKM_Tecnologia_todos.csv', path_vkm = r'VKM_Tecnologia_todos.csv',transantiago_aux = r'Aux_transantiago/'):
    regiones_dicc = {'R1':'I', 'R2':'II','R3':'III','R4':'IV','R5':'V','R6':'VI','R7':'VII',
                     'R8':'VIII','R9':'IX','R10':'X','R11':'XI','R12':'XII','RM':'RM','R14':'XIV','R15':'XV','R16':'XVI'}
    
    pkm_df = pd.read_csv(path_fold+path_pkm)
    tkm_df = pd.read_csv(path_fold+path_tkm)
    vkm_df = pd.read_csv(path_fold+path_vkm)
    transantiago_aux_df = pd.read_csv(path_fold + transantiago_aux+ 'Fracciones Rigidos.csv')
    
    '''Condiciones para filtrar dataframes'''
    años_cond_pkm = (pkm_df['Año'] <= 2050) &(pkm_df['Año'] >= 2017)
    urbano_pkm = (pkm_df['Ambito'] == 'urbano')
    interurbano_pkm = (pkm_df['Ambito'] == 'interurbano')

    años_cond_tkm = (tkm_df['Año'] <= 2050) &(tkm_df['Año'] >= 2017)
    urbano_tkm = (tkm_df['Ambito'] == 'urbano')
    interurbano_tkm = (tkm_df['Ambito'] == 'interurbano')

    años_cond_vkm = (vkm_df['Año'] <= 2050) &(vkm_df['Año'] >= 2017)
    interurbano_vkm = (vkm_df['Ambito'] == 'interurbano')
    urbano_vkm = (vkm_df['Ambito'] == 'urbano')
    regiones = (vkm_df['Region'] != 'RM')
    RM = (vkm_df['Region'] == 'RM')

    ferroviario_pkm =  (pkm_df['Modo'] == 'Tren_urbano')
    ferroviario_tkm = (tkm_df['Modo'] == 'Ferroviario')
    maritimo = (tkm_df['Tipo'] == 'Carga-Maritimo')
    petroleo_comb = (tkm_df['Motorizacion'] == 'Petróleo Combustible')
    Aereo = (pkm_df['Tipo']=='Pasajero-Aereo' )
    no_internacional = (pkm_df['Ambito']!= 'internacional')
    no_GRL = (tkm_df['Modo']!= 'GRL')


    moto =  (vkm_df['Modo'] == 'Motocicleta')
    tractocamion = (vkm_df['Modo'] == 'Camión pesado')
    camion = (vkm_df['Modo'] == 'Camión liviano') | (vkm_df['Modo'] == 'Camión mediano') #Se debe incluir medianos?
    bus = (vkm_df['Modo'] == 'Bus Rígido')
    transantiago = (vkm_df['Modo'] == 'Bus Articulado')
    taxi = (vkm_df['Modo'] == 'Taxi')
    vehiculos = (vkm_df['Modo'] == 'Vehículo Liviano Pasajero')
    medianos = (vkm_df['Modo'] == 'Vehículo Liviano Comercial')

    col_pkm = ['Año','Region','Modo','Ambito','PKM_Tecnologia']
    col_vkm = ['Año','Region','Modo','Ambito','VKM_Tecnologia']
    col_aereo =['Tipo','Año','Region','Modo','Ambito','PKM_Tecnologia']
    col_tkm = ['Año','Region','Modo','Ambito','TKM_Tecnologia']
    col_mar = ['Tipo','Año','Region','Modo','Ambito','TKM_Tecnologia']
    
    
    '''Calculamos los pkm, tkm y km (vkm)'''
    tren_urb = pkm_df[col_pkm][años_cond_pkm & urbano_pkm & ferroviario_pkm]
    tren_inter = pkm_df[col_pkm][años_cond_pkm & interurbano_pkm & ferroviario_pkm]
    ferro_tkm = tkm_df[col_tkm][años_cond_tkm & no_internacional & ferroviario_tkm]

    aereo_pkm = pkm_df[col_aereo][años_cond_pkm & no_internacional & Aereo]

    mar_tkm = tkm_df[col_mar][años_cond_tkm & interurbano_tkm & maritimo & petroleo_comb]

    moto_vkm = vkm_df[col_vkm][años_cond_vkm & moto]

    tracto_vkm = vkm_df[col_vkm][años_cond_vkm & tractocamion]
    camion_vkm = vkm_df[col_vkm][años_cond_vkm & camion]

    bus_inter_reg = vkm_df[col_vkm][años_cond_vkm & bus & interurbano_vkm & regiones]
    bus_urb_reg = vkm_df[col_vkm][años_cond_vkm & bus & urbano_vkm & regiones]
    bus_inter_rm = vkm_df[col_vkm][años_cond_vkm & bus & interurbano_vkm & RM]
    bus_urb_rm = vkm_df[col_vkm][años_cond_vkm & bus & urbano_vkm & RM]
    transantiago_vkm = vkm_df[col_vkm][años_cond_vkm & transantiago]

    taxi_vkm = vkm_df[col_vkm][años_cond_vkm & taxi]
    veh_vkm = vkm_df[col_vkm][años_cond_vkm & vehiculos]
    med_vkm = vkm_df[col_vkm][años_cond_vkm & medianos]

    dfs = [[veh_vkm,'km_livianos'],[taxi_vkm,'km_taxi'],[transantiago_vkm,'km_transantiago'],[bus_urb_rm, 'km_bus_urbano_RM'],[bus_inter_rm, 'km_bus_interurbano_RM'],[bus_urb_reg, 'km_bus_urbano'],[bus_inter_reg, 'km_bus_interurbano'],[med_vkm,'km_medianos'],[camion_vkm,'km_camion'],[tracto_vkm,'km_tractocamion'],[moto_vkm,'km_moto'],[aereo_pkm, 'pkm_aereo'],[mar_tkm,'tkm_maritimo'],[tren_urb,'pkm_ferroviario_urb'],[tren_inter,'pkm_ferroviario_inter'],[ferro_tkm, 'tkm_ferroviario']]

    df_list = []
    for sets in dfs:
        if sets[1] == 'pkm_aereo' or sets[1] == 'tkm_maritimo':
            sets[0].drop(columns=['Tipo'],inplace=True)
        sets[0].drop(columns=['Modo','Ambito'],inplace=True)
        sets[0] = sets[0].groupby(['Año','Region']).sum()
        sets[0].reset_index(inplace=True)
        sets[0].columns = ['Año','Region',sets[1]]
        sets[0][sets[1]] = sets[0][sets[1]]/1000
        df_list.append(sets[0])

    
    '''Unimos los dataframes'''
    df_merged = reduce(lambda  left,right: pd.merge(left,right,on=['Año','Region'],how='outer'), df_list)
    #df_merged = df_merged.astype(int)
    df_merged.set_index('Año',inplace=True)
    #Separar miles utilzando ","
    #df_merged.loc[:,df_merged.columns!="Año"] = df_merged.loc[:,df_merged.columns!="Año"].applymap('{:,}'.format)

    df_merged['Region'] = df_merged['Region'].map(regiones_dicc)
    df_pivoted = df_merged.reset_index().pivot_table(index='Año', columns='Region', values=df_merged.columns[1:])
    column_template = '{}_{}'

    # Crea una lista con los nuevos nombres de las columnas
    new_columns = []
    for col, region in df_pivoted.columns:
        new_columns.append(column_template.format(col, region))

    # Renombra las columnas utilizando la lista de nuevos nombres
    df_pivoted.columns = new_columns
    
    #format_template = '{:,.0f}'

    # Aplica la plantilla de formato a todas las celdas del DataFrame
    #df_pivoted = df_pivoted.applymap(lambda x: format_template.format(x) if isinstance(x, float) else x)

    cols_mod = []
    for elemento in df_pivoted.columns:
        if elemento.endswith("_RM_RM"):
            cols_mod.append(elemento[:-3])
        else:
            if elemento == 'km_transantiago_RM':
                cols_mod.append('km_transantiago')
            else:
                cols_mod.append(elemento)
    df_pivoted.columns = cols_mod 

   #ordenar
    col_ord = ['km_livianos','km_taxi','km_transantiago','km_bus_urbano','km_bus_interurbano',
               'km_medianos','km_camion','km_tractocamion','km_moto','pkm_aereo',
               'tkm_maritimo','pkm_ferroviario_urb','pkm_ferroviario_inter','tkm_ferroviario']
    reg_ord = ['XV','I','II','III','IV','V','RM','VI','VII','XVI','VIII','IX',
                   'XIV','X','XI','XII']
    veh_ord = []
    for c in col_ord:
        if c == 'km_transantiago':
            veh_ord.append(c)
        else:
            for r in reg_ord:
                veh_ord.append(c+'_'+r)
    df_pivoted = df_pivoted[veh_ord]

    #Ajustamos transantiago y buses rigidos rm urbano
    df_pivoted.reset_index(inplace=True)
    df_pivoted = df_pivoted.merge(transantiago_aux_df,how='left')

    df_pivoted['km_transantiago'] = df_pivoted['km_transantiago'] + df_pivoted['km_bus_urbano_RM']* df_pivoted['Fraccion Rigidos']
    df_pivoted['km_bus_urbano_RM'] = df_pivoted['km_bus_urbano_RM']*(1-df_pivoted['Fraccion Rigidos'])
    df_pivoted.drop(columns='Fraccion Rigidos',inplace=True)
    df_pivoted.set_index('Año',inplace=True)
    return df_pivoted




def demanda_regional_sectores_pib(path):
    
    #Cargamos los datos del PIB Regional y su respectivo factor por sector
    factor_pib = pd.read_excel(path,sheet_name='datos_factor_PIB')
    factor_pib = cambiar_estructura(factor_pib,'factor').reset_index()

    pib_regional= pd.read_excel(path,sheet_name='datos_PIB_regional')
    pib_regional=cambiar_estructura(pib_regional,'pib').reset_index()
    
    salitre_df = pd.read_excel(path,sheet_name='aux_salitre')
    salitre_df = cambiar_estructura(salitre_df,'valor_salitre').reset_index()

    datos_entrada = pd.read_excel(path,sheet_name='datos_entrada').dropna(subset=[2017]).drop(columns='Fuente')
    datos_entrada = datos_entrada.rename(columns={'Variable': 'variable'})
    datos_entrada = cambiar_estructura(datos_entrada,'demanda').reset_index()

    #Juntamos todo en un solo dataframe
    df_regional = factor_pib.merge(pib_regional,how='left',on=['Region','Año'])
    df_regional = df_regional.merge(salitre_df,how='left',on=['Region','Año','variable'])
    df_regional = df_regional.merge(datos_entrada,how='left',on= ['variable','Año'])
    df_regional.set_index(['Region','Año','variable'],inplace=True)
    df_regional['demanda_regional'] = df_regional['demanda']*df_regional['factor']
    #Comenzamos los cálculos de forma análoga a nivel nacional pero multiplicando el pib por su respectivo factor
    #Primero definimos los parámetros econométricos
    global alfa_com,beta_com,a_cem,b_cem,c_cem,d_cem,a_mineral,b_mineral,a_manuf,b_manuf,c_manuf

    #Comercial
    alfa_com = param_e.loc['alfa']['Valor']
    beta_com = param_e.loc['beta']['Valor']

    #Cemento
    a_cem = datos_cemento.loc['a']['Valor']
    b_cem = datos_cemento.loc['b']['Valor']
    c_cem = datos_cemento.loc['c']['Valor']
    d_cem = datos_cemento.loc['d']['Valor']

    #Mineral
    a_mineral = datos_mineral.loc['a']['Valor']
    b_mineral = datos_mineral.loc['b']['Valor']

    #Manufactura
    a_manuf = datos_manufactura.loc['a']['Valor']
    b_manuf = datos_manufactura.loc['b']['Valor']
    c_manuf = datos_manufactura.loc['c']['Valor']

    #Energia Publico no requiere parametros adicionales
    #Ahora creamos los elementos a recorrer (años y region)
    años = [i for i in range(2017,2051)]
    regiones = ['I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII','RM','XIV','XV','XVI']
    sectores = ['energia_comercial','cemento','mineral','manufactura','energia_publico']

    df_regional.to_csv('prueba_regional.csv',encoding='latin1') #Esto esta en construccion
    #Comenzamos el calculo
    for reg in regiones:
        for año in años:
            #Comercial
            pib = df_regional.loc[(reg,año,'energia_comercial')]['pib']
            comercial_v = df_regional.loc[(reg,año,'energia_comercial')]['demanda_regional']
            factor_com = df_regional.loc[(reg,año,'energia_comercial')]['factor']
            if pd.isna(comercial_v):
                df_regional.loc[(reg,año,'energia_comercial'),'demanda_regional'] = np.exp(alfa_com+beta_com*np.log(pib*factor_com))
            
            #Cemento
            pib = df_regional.loc[(reg,año,'cemento')]['pib']
            factor_cemento = df_regional.loc[(reg,año,'cemento')]['factor']

            cemento_v = df_regional.loc[(reg,año,'cemento')]['demanda_regional']
            if pd.isna(cemento_v):
                pib_anterior = df_regional.loc[(reg,año-1,'cemento')]['pib']
                cemento_anterior = df_regional.loc[(reg,año-1,'cemento')]['demanda_regional']
                
                df_regional.loc[(reg,año,'cemento'),'demanda_regional'] = np.exp(a_cem + np.log(cemento_anterior)*b_cem + np.log(pib*factor_cemento)*c_cem+np.log(pib_anterior*factor_cemento)*d_cem)
            
            #Mineral
            pib = df_regional.loc[(reg,año,'mineral')]['pib']
            factor_mineral = df_regional.loc[(reg,año,'mineral')]['factor']
            salitre = df_regional.loc[(reg,año,'salitre')]['valor_salitre']

            mineral_v = df_regional.loc[(reg,año,'mineral')]['demanda_regional']
            if pd.isna(mineral_v):
                df_regional.loc[(reg,año,'mineral'),'demanda_regional'] = np.exp(np.log(pib*factor_mineral)*b_mineral + a_mineral) - salitre
            
            #Manufactura
            pib = df_regional.loc[(reg,año,'manufactura')]['pib']
            factor_manufactura = df_regional.loc[(reg,año,'manufactura')]['factor']

            manufactura_v = df_regional.loc[(reg,año,'manufactura')]['demanda_regional']
            if pd.isna(manufactura_v):
                manufactura_anterior = df_regional.loc[(reg,año-1,'manufactura')]['demanda_regional']

                df_regional.loc[(reg,año,'manufactura'),'demanda_regional'] = np.exp(np.log(pib*factor_manufactura)*c_manuf + b_manuf*np.log(manufactura_anterior)+a_manuf)
            
            #Energia publico
            pib_hist = df_regional.loc[(reg,2017,'energia_publico')]['pib']
            energia_publico_hist = df_regional.loc[(reg,2017,'energia_publico')]['demanda_regional']
            pib = df_regional.loc[(reg,año,'energia_publico')]['pib']

            energia_publico_v = df_regional.loc[(reg,año,'energia_publico')]['demanda_regional']
            if pd.isna(energia_publico_v):
                df_regional.loc[(reg,año,'energia_publico'),'demanda_regional'] = (pib/pib_hist)*energia_publico_hist

            for sec in sectores:
                df_regional.loc[(reg,año,sec),'historico'] = df_regional.loc[(reg,2017,sec)]['demanda_regional']
    
    df_regional['existente'] = df_regional[['historico','demanda_regional']].min(axis=1)
    df_regional['nuevo'] = df_regional['demanda_regional'] - df_regional['existente']

    df_regional = df_regional[['demanda_regional','existente','nuevo']]
    df_regional = df_regional.reset_index()

    wb = openpyxl.load_workbook(path)

 # Crear una nueva hoja
    sheet_name = 'datos_demanda_regional'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    color = Color(rgb='00B0F0')
    sheet.sheet_properties.tabColor = color

    # Escribir encabezados en primera fila
    for j, header in enumerate(df_regional.columns, start=1):
        sheet.cell(row=1, column=j, value=header)

    # Escribir el contenido del DataFrame en la nueva hoja
    for i, row in df_regional.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    wb.save(path)
    wb.close()


def crear_proyeccion_demanda(path,obs):
    df = calcular_proyecciones(path,obs)

    if 'km_recorridos' in df.columns:
        df.rename(columns={'km_recorridos': 'km_livianos'}, inplace=True)
    if 'km_recorridos_taxi' in df.columns:
        df.rename(columns={'km_recorridos_taxi': 'km_taxi'}, inplace=True)

    for dir in [path,obs]:
        wb = openpyxl.load_workbook(dir)
        datoss = df.copy()
        datoss = datoss.melt(id_vars=["Año"], value_vars=datoss.columns.difference(["Año"]), var_name="Variable", value_name="Valor").set_index(["Año","Variable"])
        datoss["Valor"]=datoss.loc[:,datoss.columns[0]]

        unidades_dic = {'PIB':'miles mill CLP','PIB (mill USD)': 'mill USD', 'tasa PIB': '%', 'Población': 'habitantes', 
                        'Viviendas': 'viviendas', 'HAB/viv': 'habitantes/vivienda', 'PIB per cápita (usd/hab)': 'usd/habitantes',
                        'electricidad_res':None, 'calefaccion': None, 'ACS':None, 'coccion':None, 'energia_comercial': 'Tcal', 'manufactura': 'Tcal',
                        'celulosa': 'miles ton', 'azucar':'miles ton', 'cemento':'Tcal', 'hierro':'Tcal', 'pesca':'miles ton', 'acero':'miles ton', 'petroquimica':'miles ton', 
                        'mineral': 'Tcal', 'cobre': 'miles ton', 'energia_publico': 'Tcal', 
                        'km_livianos': 'miles km', 'km_taxi': 'miles km', 'km_transantiago': 'miles km', 'km_bus_urbano_RM': 'miles km', 'km_bus_interurbano_RM':'miles km',
                        'km_bus_urbano_regiones': 'miles km', 'km_bus_interurbano_regiones': 'miles km', 'km_medianos': 'miles km', 'km_camion': 'miles km', 
                        'km_tractocamion': 'miles km', 'km_moto': 'miles km', 'pkm_aereo': 'miles pkm', 'tkm_maritimo': 'miles tkm', 'pkm_ferroviario_urb': 'miles pkm', 'pkm_ferroviario_inter': 'miles pkm',
                        'tkm_ferroviario': 'miles tkm', 'manufactura_existente': 'Tcal', 'celulosa_existente' :'miles ton', 'azucar_existente': 'miles ton', 'cemento_existente': 'Tcal',
                        'hierro_existente': 'Tcal', 'pesca_existente': 'miles ton', 'acero_existente': 'miles ton', 'petroquimica_existente': 'miles ton' , 'mineral_existente': 'Tcal', 'cobre_existente': 'miles ton',
                        'manufactura_nuevo': 'Tcal', 'celulosa_nuevo' :'miles ton', 'azucar_nuevo': 'miles ton', 'cemento_nuevo': 'Tcal','hierro_nuevo': 'Tcal', 'pesca_nuevo': 'miles ton', 
                        'acero_nuevo': 'miles ton', 'petroquimica_nuevo': 'miles ton' , 'mineral_nuevo': 'Tcal', 'cobre_nuevo': 'miles ton','calefaccion_biomasa_COM': 'Tcal','coccion_biomasa_COM':'Tcal',
                        'calefaccion_gas_licuado_COM':'Tcal', 'coccion_gas_licuado_COM':'Tcal', 'ACS_gas_licuado_COM': 'Tcal', 'calefaccion_gas_natural_COM':'Tcal', 'coccion_gas_natural_COM': 'Tcal', 'ACS_gas_natural_COM':'Tcal',
                        'calefaccion_kerosene_COM':'Tcal','calefaccion_electricidad_COM':'Tcal','coccion_electricidad_COM':'Tcal','refrigeracion_COM':'Tcal','iluminacion_COM':'Tcal','otros_electricos_COM':'Tcal',
                        'enfriamiento_COM':'Tcal','ACS_electricidad_COM':'Tcal','electrogeno_diesel_COM':'Tcal','transporte_diesel_COM':'Tcal', 'Tipo cambio dólar': 'CLP/USD', 'Tipo cambio euro': 'CLP/EUR','Tasa Población':
                        '%'}
                

        datoss.reset_index(inplace=True)
        datoss["Unidad"] = datoss["Variable"].map(unidades_dic)
        datoss = datoss[['Año','Variable','Unidad','Valor']]
        datoss= datoss.sort_values('Año')

        # Crear una nueva hoja
        sheet_name = 'datos_proyeccion_demanda'
        if sheet_name in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
        sheet = wb.create_sheet(title=sheet_name)
        color = Color(rgb='00B0F0')
        sheet.sheet_properties.tabColor = color

        # Escribir encabezados en primera fila
        for j, header in enumerate(datoss.columns, start=1):
            sheet.cell(row=1, column=j, value=header)

        # Escribir el contenido del DataFrame en la nueva hoja
        for i, row in datoss.iterrows():
            for j, value in enumerate(row, start=1):
                sheet.cell(row=i+2, column=j, value=value)

        wb.save(dir)
        wb.close()



#Funcion para actualizar proyecciones en modelo PMR
def actualizar_proyecciones_demanda(path,obs):
    
    
    # Abre el archivo de Excel
    print('Cargando archivo de modelo de datos')
    #separador = os.path.sep
    #dir_actual = os.path.dirname(os.path.abspath('proyecciones_demanda.py'))
    #dir = separador.join(dir_actual.split(separador)[:-1])
    #path = dir + obs

    df = calcular_proyecciones(path,obs)
    crear_proyeccion_demanda(path,obs)
    demanda_regional_sectores_pib(path)
    new_table = pd.read_excel(path,sheet_name='datos_proyeccion_demanda')
    largo_tabla = str(len(new_table)+1)

    wb = openpyxl.load_workbook(obs)
    
    # Selecciona la hoja de trabajo
    ws = wb['Proyecciones demanda']
    wc = wb['Demanda_Cobre']
    df_cobre = df.reset_index()[['Año','cobre']]

    fila_excel = 2  # La fila 1 corresponde a la fila de encabezados
    bar = Bar('Actualizando filas:', max=len(df.index))
    for fila in df.index:
        time.sleep(0.2)
        for col in df.columns:
            col_n = list(df.columns).index(col) + 1
            celda = ws.cell(fila_excel, col_n)
            if col != 'Año':
                if col_n == len(df.columns):
                    pass
                else:
                    fila_ind= str(celda.row)
                    col_letra = str(celda.column_letter)
                    #formula = "=SUMIFS(datos_proyeccion_demanda!$D$2:$D$2687,datos_proyeccion_demanda!$A$2:$A$2687,'Proyecciones demanda'!$A" + fila_ind + ",datos_proyeccion_demanda!$B$2:$B$2687,'Proyecciones demanda'!" + col_letra + "$1)"
                    formula = "=SUMIFS(datos_proyeccion_demanda!$D$2:$D$"+largo_tabla+",datos_proyeccion_demanda!$A$2:$A$"+largo_tabla+",'Proyecciones demanda'!$A" + fila_ind + ",datos_proyeccion_demanda!$B$2:$B$"+largo_tabla+",'Proyecciones demanda'!" + col_letra + "$1)"
                    celda.value = formula
            if col == 'Año':
                celda.value = df.at[fila, col]
        fila_excel += 1
        bar.next()
    bar.finish()

    ws.freeze_panes=ws['B2']

    fila_excel_2= 2
    bar2 = Bar('Actualizando demanda cobre:', max=len(df_cobre.index))
    for fila in df_cobre.index:
        time.sleep(0.2)
        for col in df_cobre.columns:
            if col != 'Año':
                celda = wc.cell(fila_excel_2, list(df_cobre.columns).index(col) + 1) 
                fila_ind2= str(celda.row)
                celda.value = '=SUMIFS(datos_proyeccion_demanda!$D$2:$D$'+largo_tabla+',datos_proyeccion_demanda!$A$2:$A$'+largo_tabla+',Demanda_Cobre!A'+ fila_ind2+',datos_proyeccion_demanda!$B$2:$B$'+largo_tabla+',"cobre")'
        fila_excel_2 += 1
        bar2.next()
    bar2.finish()


    # Guarda los cambios en el archivo de Excel con xlwings
    print('Guardando resultados')
    wb.save(obs)
    wb.close()


    #Calculo de proyecciones transporte regional
    df_trans_reg = proyecciones_step_regiones()
    df_hist_transporte_reg, año_hist_transporte_reg = obtener_año_hist_reg(path,'datos_historicos_regiones')
    cols_historicas = df_hist_transporte_reg.columns
    print(f'Agregando información histórica de transporte regional')
    for col in cols_historicas:
        for año in list(df_trans_reg.index.values):
            if año <= año_hist_transporte_reg[col]:
                transporte_historico = df_hist_transporte_reg.loc[año][col]
                if not pd.isna(transporte_historico):
                    df_trans_reg.loc[año,col] = transporte_historico

    df_trans_reg = df_trans_reg.reset_index()
    wb = openpyxl.load_workbook(obs)

    # Crear una nueva hoja
    sheet_name = 'data_transporte_regional'
    if sheet_name in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    sheet = wb.create_sheet(title=sheet_name)
    color = Color(rgb='C00000')
    sheet.sheet_properties.tabColor = color

    # Escribir encabezados en primera fila
    for j, header in enumerate(df_trans_reg.columns, start=1):
        sheet.cell(row=1, column=j, value=header)

    # Escribir el contenido del DataFrame en la nueva hoja
    for i, row in df_trans_reg.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=value)

    wb.save(obs)
    wb.close()


    print('Generando datos de salida')
    generar_salidas(path)
    
    print('Recalculando fórmulas en el libro Excel')
    app = xw.App(visible=False, add_book=False) # se utiliza visible=False para que Excel no se muestre
    app.display_alerts = False
    book = app.books.open(obs)
    book2 = app.books.open(path)
    app.calculate()
    book.save()
    book2.save()
    os.chmod(obs, 0o777)
    os.chmod(path,0o777)
    book.close()
    book2.close()
    app.quit()






#Ejecutamos el calculo
def main2(path,obs):
    start_time = time.time()
    print()
    print('Actualización de proyección de demanda')
    actualizar_proyecciones_demanda(path,obs)
    print()
    print('Archivo modificado')
    print()
    print("--- %s seconds ---" % (time.time() - start_time))


def subir_datos():
    print('Por favor ingrese caso IPMR')
# Create a tkinter window
    root = tk.Tk()
    root.withdraw()  # Hide the main window

# Open a file dialog and get the selected file's name
    file_path = filedialog.askopenfilename()
    #file_name = file_path.split('/')[-1]
    print(f'Archivo cargado: {file_path.split("/")[-1]}')
    return file_path


if __name__ == '__main__':
    directory = "datos_proyecciones"
    if not os.path.exists(directory):
        os.mkdir(directory)
    main(ruta=os.sep.join(['datos_proyecciones']),
         caso_base=r'Modelo_demanda_IPMR.xlsx')
    main2(path = r'Modelo_demanda_IPMR.xlsx', obs =subir_datos() )
